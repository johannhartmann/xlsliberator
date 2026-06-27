#!/usr/bin/env python3
"""Generate the demo sample workbooks shipped with the web app.

These are genuine ``.xlsx`` files (multi-sheet, cross-sheet formulas) so the
landing-page demo performs a real conversion rather than a simulation. Run
from the repository root::

    python scripts/generate_web_samples.py

Output lands in ``src/xlsliberator/web/static/samples/``.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SAMPLES_DIR = Path(__file__).resolve().parent.parent / "src/xlsliberator/web/static/samples"

HEADER_FILL = PatternFill("solid", fgColor="15315C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="15315C")


def _header(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_haushalt() -> Workbook:
    """A municipal-style budget workbook with cross-sheet rollups."""
    wb = Workbook()

    months = [
        "Januar",
        "Februar",
        "März",
        "April",
        "Mai",
        "Juni",
        "Juli",
        "August",
        "September",
        "Oktober",
        "November",
        "Dezember",
    ]
    income_cats = ["Steuern", "Gebühren", "Zuweisungen", "Sonstige Erträge"]
    expense_cats = [
        "Personal",
        "Sachaufwand",
        "Investitionen",
        "Zinsen",
        "Transferleistungen",
        "Unterhalt",
    ]

    einnahmen = wb.active
    einnahmen.title = "Einnahmen"
    einnahmen.cell(row=1, column=1, value="Einnahmen 2025 (in €)").font = TITLE_FONT
    _header(einnahmen, 2, ["Kategorie", *months, "Summe"])
    for i, cat in enumerate(income_cats):
        row = 3 + i
        einnahmen.cell(row=row, column=1, value=cat)
        for m in range(12):
            base = 12000 + i * 4200 + m * 130
            einnahmen.cell(row=row, column=2 + m, value=base)
        last = 2 + 12
        einnahmen.cell(row=row, column=last + 1, value=f"=SUM(B{row}:M{row})")
    total_row = 3 + len(income_cats)
    einnahmen.cell(row=total_row, column=1, value="Gesamt").font = Font(bold=True)
    for m in range(12):
        col = 2 + m
        letter = einnahmen.cell(row=3, column=col).column_letter
        einnahmen.cell(row=total_row, column=col, value=f"=SUM({letter}3:{letter}{total_row - 1})")
    einnahmen.cell(row=total_row, column=14, value=f"=SUM(N3:N{total_row - 1})")

    ausgaben = wb.create_sheet("Ausgaben")
    ausgaben.cell(row=1, column=1, value="Ausgaben 2025 (in €)").font = TITLE_FONT
    _header(ausgaben, 2, ["Kategorie", *months, "Summe"])
    for i, cat in enumerate(expense_cats):
        row = 3 + i
        ausgaben.cell(row=row, column=1, value=cat)
        for m in range(12):
            base = 9000 + i * 3100 + m * 95
            ausgaben.cell(row=row, column=2 + m, value=base)
        ausgaben.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
    exp_total_row = 3 + len(expense_cats)
    ausgaben.cell(row=exp_total_row, column=1, value="Gesamt").font = Font(bold=True)
    for m in range(12):
        col = 2 + m
        letter = ausgaben.cell(row=3, column=col).column_letter
        ausgaben.cell(
            row=exp_total_row, column=col, value=f"=SUM({letter}3:{letter}{exp_total_row - 1})"
        )
    ausgaben.cell(row=exp_total_row, column=14, value=f"=SUM(N3:N{exp_total_row - 1})")

    bilanz = wb.create_sheet("Bilanz")
    bilanz.cell(row=1, column=1, value="Bilanz 2025").font = TITLE_FONT
    _header(bilanz, 2, ["Position", "Betrag (€)"])
    bilanz.cell(row=3, column=1, value="Einnahmen gesamt")
    bilanz.cell(row=3, column=2, value=f"=Einnahmen.N{total_row}")
    bilanz.cell(row=4, column=1, value="Ausgaben gesamt")
    bilanz.cell(row=4, column=2, value=f"=Ausgaben.N{exp_total_row}")
    bilanz.cell(row=5, column=1, value="Saldo").font = Font(bold=True)
    bilanz.cell(row=5, column=2, value="=B3-B4").font = Font(bold=True)
    bilanz.cell(row=6, column=1, value="Deckungsquote")
    bilanz.cell(row=6, column=2, value="=ROUND(B3/B4*100,2)")
    bilanz.cell(row=7, column=1, value="Status")
    bilanz.cell(row=7, column=2, value='=IF(B5>=0,"ausgeglichen","Defizit")')

    uebersicht = wb.create_sheet("Übersicht", 0)
    uebersicht.cell(row=1, column=1, value="Haushaltsübersicht 2025").font = TITLE_FONT
    _header(uebersicht, 2, ["Kennzahl", "Wert"])
    rows = [
        ("Einnahmen gesamt (€)", f"=Einnahmen.N{total_row}"),
        ("Ausgaben gesamt (€)", f"=Ausgaben.N{exp_total_row}"),
        ("Saldo (€)", "=Bilanz.B5"),
        ("Deckungsquote (%)", "=Bilanz.B6"),
        ("Ø Einnahmen/Monat (€)", f"=ROUND(Einnahmen.N{total_row}/12,2)"),
        ("Ø Ausgaben/Monat (€)", f"=ROUND(Ausgaben.N{exp_total_row}/12,2)"),
    ]
    for i, (label, formula) in enumerate(rows):
        uebersicht.cell(row=3 + i, column=1, value=label)
        uebersicht.cell(row=3 + i, column=2, value=formula)

    return wb


def build_kennzahlen() -> Workbook:
    """A quarterly KPI workbook, formulas only, no macros."""
    wb = Workbook()

    umsatz = wb.active
    umsatz.title = "Umsatz"
    umsatz.cell(row=1, column=1, value="Umsatz Q2 (in T€)").font = TITLE_FONT
    _header(umsatz, 2, ["Produkt", "April", "Mai", "Juni", "Q2 gesamt", "Anteil %"])
    products = ["Software", "Beratung", "Wartung", "Schulung", "Lizenzen"]
    for i, prod in enumerate(products):
        row = 3 + i
        umsatz.cell(row=row, column=1, value=prod)
        for m in range(3):
            umsatz.cell(row=row, column=2 + m, value=120 + i * 45 + m * 18)
        umsatz.cell(row=row, column=5, value=f"=SUM(B{row}:D{row})")
        umsatz.cell(
            row=row, column=6, value=f"=ROUND(E{row}/SUM($E$3:$E${2 + len(products)})*100,1)"
        )
    tot = 3 + len(products)
    umsatz.cell(row=tot, column=1, value="Gesamt").font = Font(bold=True)
    for col in range(2, 6):
        letter = umsatz.cell(row=3, column=col).column_letter
        umsatz.cell(row=tot, column=col, value=f"=SUM({letter}3:{letter}{tot - 1})")

    trend = wb.create_sheet("Trend")
    trend.cell(row=1, column=1, value="Monatstrend").font = TITLE_FONT
    _header(trend, 2, ["Monat", "Umsatz", "Veränderung %"])
    trend.cell(row=3, column=1, value="April")
    trend.cell(row=3, column=2, value="=Umsatz.B8")
    trend.cell(row=4, column=1, value="Mai")
    trend.cell(row=4, column=2, value="=Umsatz.C8")
    trend.cell(row=4, column=3, value="=ROUND((B4-B3)/B3*100,1)")
    trend.cell(row=5, column=1, value="Juni")
    trend.cell(row=5, column=2, value="=Umsatz.D8")
    trend.cell(row=5, column=3, value="=ROUND((B5-B4)/B4*100,1)")

    kpis = wb.create_sheet("KPIs", 0)
    kpis.cell(row=1, column=1, value="Kennzahlen Q2").font = TITLE_FONT
    _header(kpis, 2, ["Kennzahl", "Wert"])
    kpi_rows = [
        ("Umsatz gesamt (T€)", "=Umsatz.E8"),
        ("Top-Produkt-Anteil (%)", "=MAX(Umsatz.F3:F7)"),
        ("Ø Umsatz/Produkt (T€)", "=ROUND(Umsatz.E8/5,1)"),
        ("Wachstum Mai→Juni (%)", "=Trend.C5"),
    ]
    for i, (label, formula) in enumerate(kpi_rows):
        kpis.cell(row=3 + i, column=1, value=label)
        kpis.cell(row=3 + i, column=2, value=formula)

    return wb


def main() -> None:
    SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
    build_haushalt().save(SAMPLES_DIR / "Haushalt_2025.xlsx")
    build_kennzahlen().save(SAMPLES_DIR / "Kennzahlen_Q2.xlsx")
    for path in sorted(SAMPLES_DIR.glob("*.xlsx")):
        print(f"wrote {path} ({path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
