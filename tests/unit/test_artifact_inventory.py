"""Canonical artifact inventory and loss-accounting tests."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from xlsliberator.artifact_inventory import (
    diff_inventories,
    disposition_coverage_errors,
    inventory_ods,
)
from xlsliberator.inspect_workbook import inspect_workbook
from xlsliberator.ir_models import ExtractionStats, WorkbookIR
from xlsliberator.validation_models import (
    ArtifactCoverage,
    ArtifactDisposition,
    ArtifactDispositionKind,
    CanonicalArtifactIR,
    SourceRef,
    WorkbookArtifactIR,
)


def _rich_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    workbook.properties.title = "Artifact fixture"
    sheet: Worksheet = workbook.worksheets[0]
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["A", 2])
    sheet.append(["B", 3])
    sheet["B4"] = "=SUM(B2:B3)"
    sheet["A2"].comment = Comment("comment", "tester")
    sheet["A3"].hyperlink = "https://example.invalid/item"
    sheet["B2"].number_format = "0.00"
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    sheet.row_dimensions[2].height = 24
    sheet.column_dimensions["A"].width = 20
    sheet.merge_cells("C1:D1")
    sheet["C1"] = "Merged"
    table = Table(displayName="DataTable", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    sheet.conditional_formatting.add("B2:B4", CellIsRule(operator="greaterThan", formula=["2"]))
    validation = DataValidation(type="whole", operator="between", formula1="0", formula2="10")
    validation.add("B2:B4")
    sheet.add_data_validation(validation)
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    sheet.add_chart(chart, "F2")
    sheet.print_area = "A1:D4"
    sheet.print_title_rows = "1:1"
    sheet.page_setup.orientation = "landscape"
    workbook.save(path)
    workbook.close()
    injected = {
        "xl/pivotTables/pivotTable1.xml": b"<pivotTableDefinition/>",
        "xl/pivotCache/pivotCacheDefinition1.bin": b"pivot-cache",
        "xl/slicers/slicer1.xml": b"<slicer/>",
        "xl/activeX/activeX1.bin": b"activex",
        "xl/ctrlProps/ctrlProp1.xml": b'<control fmlaMacro="RunMe"/>',
        "xl/queries/query1.xml": b"<query/>",
        "xl/connections.xml": b"<connections/>",
        "xl/model/model.bin": b"data-model",
        "xl/embeddings/oleObject1.bin": b"ole-object",
        "xl/media/image1.png": b"synthetic-image",
        "xl/drawings/drawing99.xml": (
            b'<xdr:wsDr xmlns:xdr="urn:xdr" xmlns:a="urn:a">'
            b"<xdr:sp><xdr:txBody><a:p/></xdr:txBody></xdr:sp></xdr:wsDr>"
        ),
        "xl/vbaProject.bin": b"synthetic-vba-project",
        "xl/unknown/customExtension.bin": b"unknown-extension",
    }
    with ZipFile(path, "a", ZIP_DEFLATED) as archive:
        for name, payload in injected.items():
            archive.writestr(name, payload)


def test_ooxml_inventory_covers_semantic_raw_and_unknown_artifact_families(
    tmp_path: Path,
) -> None:
    path = tmp_path / "rich.xlsx"
    _rich_workbook(path)

    inventory = inspect_workbook(path)

    families = {artifact.family for artifact in inventory.artifacts}
    assert {
        "workbook_metadata",
        "worksheet_metadata",
        "cell",
        "cell_semantics",
        "comment",
        "style",
        "row_dimension",
        "column_dimension",
        "merged_cell",
        "table",
        "structured_reference",
        "conditional_formatting",
        "data_validation",
        "hyperlink",
        "calculation_settings",
        "chart",
        "drawing",
        "image",
        "shape",
        "text_box",
        "pivot_table",
        "pivot_cache",
        "slicer_timeline",
        "control",
        "activex",
        "event_binding",
        "vba_project",
        "vba_project_metadata",
        "vba_reference",
        "query",
        "connection",
        "data_model",
        "embedded_ole",
        "print_area",
        "page_setup",
        "header_footer",
        "package_part",
        "relationship",
    } <= families
    unknown = [
        artifact
        for artifact in inventory.artifacts
        if artifact.raw_path == "xl/unknown/customExtension.bin"
        and artifact.family == "package_part"
    ]
    assert len(unknown) == 1
    assert unknown[0].coverage is ArtifactCoverage.RAW
    assert unknown[0].known is False
    package_files = {item.filename for item in ZipFile(path).infolist() if not item.is_dir()}
    inventoried_files = {
        artifact.raw_path for artifact in inventory.artifacts if artifact.family == "package_part"
    }
    assert inventoried_files == package_files
    assert inventory.schema_version == "3.0.0"
    assert inventory.source_sha256


def test_stable_artifact_ids_do_not_depend_on_workbook_path(tmp_path: Path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "renamed.xlsx"
    _rich_workbook(first)
    shutil.copy2(first, second)

    first_inventory = inspect_workbook(first)
    second_inventory = inspect_workbook(second)

    assert [item.artifact_id for item in first_inventory.artifacts] == [
        item.artifact_id for item in second_inventory.artifacts
    ]


def test_diff_assigns_exactly_one_disposition_and_rejects_unknown_loss(tmp_path: Path) -> None:
    source_path = tmp_path / "rich.xlsx"
    _rich_workbook(source_path)
    source = inspect_workbook(source_path)
    target = source.model_copy(deep=True)
    target.inventory_role = "target"
    difference = diff_inventories(source, target)
    source.dispositions = difference.dispositions

    assert len(difference.dispositions) == len(source.artifacts)
    assert disposition_coverage_errors(source) == []
    assert all(
        disposition.disposition is ArtifactDispositionKind.PRESERVED
        or disposition.disposition is ArtifactDispositionKind.TRANSLATED
        for disposition in difference.dispositions
    )

    target.artifacts = [
        artifact
        for artifact in target.artifacts
        if artifact.raw_path != "xl/unknown/customExtension.bin"
    ]
    failed = diff_inventories(source, target)
    source.dispositions = failed.dispositions

    errors = disposition_coverage_errors(source)
    assert any("unknown package part was not preserved" in error for error in errors)


def test_diff_does_not_treat_an_unrelated_same_family_artifact_as_translation() -> None:
    source_artifact = CanonicalArtifactIR(
        artifact_id="artifact:cell:source",
        family="cell",
        artifact_type="cell",
        locator="sheet/Source/cell/A1",
        coverage=ArtifactCoverage.SEMANTIC,
        source_ref=SourceRef(
            source_file="source.xlsx",
            artifact_type="cell",
            artifact_id="artifact:cell:source",
        ),
    )
    target_artifact = source_artifact.model_copy(
        update={
            "artifact_id": "artifact:cell:target",
            "locator": "sheet/Other/cell/Z99",
        }
    )
    source = WorkbookArtifactIR(
        workbook=WorkbookIR(file_path="source.xlsx", file_format="xlsx"),
        artifacts=[source_artifact],
    )
    target = WorkbookArtifactIR(
        inventory_role="target",
        workbook=WorkbookIR(file_path="target.ods", file_format="ods"),
        artifacts=[target_artifact],
    )

    difference = diff_inventories(source, target)

    assert difference.matched == {}
    assert difference.missing_source_artifact_ids == ["artifact:cell:source"]
    assert difference.dispositions[0].disposition is ArtifactDispositionKind.FAILED


def test_ods_target_inventory_contains_cells_and_complete_package_parts() -> None:
    path = Path(__file__).parents[1] / "fixtures" / "scenarios" / "basic.ods"

    inventory = inventory_ods(path)

    assert inventory.inventory_role == "target"
    assert any(item.locator == "sheet/Sheet1/cell/A3" for item in inventory.artifacts)
    assert any(item.family == "cell_semantics" for item in inventory.artifacts)
    assert any(item.family == "package_part" for item in inventory.artifacts)


def test_xls_adapter_records_raw_container_and_incomplete_semantics(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"not-an-ole-file")
    import xlsliberator.inspect_workbook as inspect_module

    monkeypatch.setattr(
        inspect_module,
        "extract_workbook",
        lambda _path: (WorkbookIR(file_path=str(path), file_format="xls"), ExtractionStats()),
    )
    monkeypatch.setattr(inspect_module, "extract_vba_modules", lambda _path: [])

    inventory = inspect_module.inspect_workbook(path)

    assert any(item.family == "raw_stream" for item in inventory.artifacts)
    coverage = next(item for item in inventory.artifacts if item.family == "format_coverage")
    assert coverage.coverage is ArtifactCoverage.UNPARSED
    assert inventory.metadata["canonical_inventory"]["semantic_coverage_complete"] is False


def test_xlsb_adapter_records_every_zip_part_and_incomplete_semantics(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    path = tmp_path / "binary.xlsb"
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/workbook.bin", b"binary-workbook-records")
        archive.writestr("xl/worksheets/sheet1.bin", b"binary-sheet-records")
        archive.writestr("xl/unknown/vendorExtension.bin", b"vendor-extension")

    import xlsliberator.inspect_workbook as inspect_module

    monkeypatch.setattr(
        inspect_module,
        "extract_workbook",
        lambda _path: (WorkbookIR(file_path=str(path), file_format="xlsb"), ExtractionStats()),
    )
    monkeypatch.setattr(inspect_module, "extract_vba_modules", lambda _path: [])

    inventory = inspect_module.inspect_workbook(path)

    package_parts = {
        artifact.raw_path: artifact
        for artifact in inventory.artifacts
        if artifact.family == "package_part"
    }
    assert set(package_parts) == {
        "xl/workbook.bin",
        "xl/worksheets/sheet1.bin",
        "xl/unknown/vendorExtension.bin",
    }
    assert package_parts["xl/unknown/vendorExtension.bin"].known is False
    coverage = next(item for item in inventory.artifacts if item.family == "format_coverage")
    assert coverage.coverage is ArtifactCoverage.UNPARSED
    assert coverage.semantic_data == {"complete": False, "raw_record_coverage": True}
    assert inventory.metadata["canonical_inventory"]["semantic_coverage_complete"] is False
    assert inventory.unsupported_artifacts[0].severity.value == "error"
    assert "full binary-record coverage is not claimed" in inventory.unsupported_artifacts[0].reason


def test_missing_disposition_and_unreferenced_disposition_are_rejected() -> None:
    inventory = WorkbookArtifactIR(
        workbook=WorkbookIR(file_path="book.xlsx", file_format="xlsx"),
        artifacts=[
            CanonicalArtifactIR(
                artifact_id="artifact:cell:one",
                family="cell",
                artifact_type="cell",
                locator="sheet/Sheet1/cell/A1",
                coverage=ArtifactCoverage.SEMANTIC,
                source_ref=SourceRef(
                    source_file="book.xlsx",
                    artifact_type="cell",
                    artifact_id="artifact:cell:one",
                ),
            )
        ],
    )
    assert disposition_coverage_errors(inventory) == [
        "artifact:cell:one: expected exactly one disposition, got 0"
    ]
    inventory.dispositions = [
        ArtifactDisposition(
            source_artifact_id="artifact:cell:one",
            disposition=ArtifactDispositionKind.PRESERVED,
        )
    ]
    assert disposition_coverage_errors(inventory) == [
        "artifact:cell:one: disposition has no target or evidence reference"
    ]
