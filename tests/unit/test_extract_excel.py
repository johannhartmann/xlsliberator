"""Unit tests for Excel extraction."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from xlsliberator.extract_excel import ExtractionError, extract_workbook
from xlsliberator.ir_models import CellType


def create_test_xlsx(file_path: Path, with_formulas: bool = True) -> None:
    """Create a synthetic test .xlsx file.

    Args:
        file_path: Output file path
        with_formulas: Include formula cells
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add values
    ws["A1"] = 10
    ws["A2"] = 20
    ws["B1"] = "Hello"
    ws["B2"] = "World"

    if with_formulas:
        # Add formulas
        ws["A3"] = "=A1+A2"
        ws["A4"] = "=SUM(A1:A2)"
        ws["A5"] = "=AVERAGE(A1:A2)"
        ws["B3"] = '=CONCATENATE(B1," ",B2)'
        ws["C1"] = '=IF(A1>5,"Yes","No")'
        ws["C2"] = "=VLOOKUP(A1,A1:B2,2,FALSE)"

    # Add named range
    wb.defined_names.add(
        openpyxl.workbook.defined_name.DefinedName("TestRange", attr_text="Sheet1!$A$1:$A$2")
    )

    # Add second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = 100

    if with_formulas:
        ws2["A2"] = "=A1*2"

    wb.save(file_path)
    wb.close()


def create_test_xlsx_with_table(file_path: Path) -> None:
    """Create a test .xlsx file with a table.

    Args:
        file_path: Output file path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add data for table
    ws.append(["Name", "Age", "Score"])
    ws.append(["Alice", 25, 85])
    ws.append(["Bob", 30, 92])
    ws.append(["Charlie", 28, 78])

    # Create table
    tab = openpyxl.worksheet.table.Table(displayName="TestTable", ref="A1:C4")
    ws.add_table(tab)

    # Add formula referencing table
    ws["D2"] = "=SUM(TestTable[Score])"

    wb.save(file_path)
    wb.close()


def test_extract_xlsx_basic() -> None:
    """Test basic .xlsx extraction."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"
        create_test_xlsx(file_path)

        wb_ir, stats = extract_workbook(file_path)

        # Check workbook properties
        assert wb_ir.file_format == "xlsx"
        assert wb_ir.sheet_count == 2
        assert not wb_ir.has_macros

        # Check sheets
        assert len(wb_ir.sheets) == 2
        sheet1 = wb_ir.get_sheet_by_name("Sheet1")
        assert sheet1 is not None
        assert sheet1.name == "Sheet1"

        sheet2 = wb_ir.get_sheet_by_index(1)
        assert sheet2 is not None
        assert sheet2.name == "Sheet2"

        # Check stats
        assert stats.total_cells > 0
        assert stats.total_formulas >= 7  # 6 in Sheet1, 1 in Sheet2


def test_extract_xlsx_formulas() -> None:
    """Test formula extraction (Gate G3 requirement: ≥99% formulas)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"
        create_test_xlsx(file_path, with_formulas=True)

        wb_ir, stats = extract_workbook(file_path)

        # Find formula cells
        sheet1 = wb_ir.get_sheet_by_name("Sheet1")
        assert sheet1 is not None

        formula_cells = [c for c in sheet1.cells if c.cell_type == CellType.FORMULA]

        # Should have extracted formulas
        assert len(formula_cells) >= 6

        # Check specific formulas
        formulas = {c.address: c.formula for c in formula_cells}
        assert "A3" in formulas
        assert formulas["A3"] == "=A1+A2"

        # Check extraction rate (Gate G3)
        assert stats.formula_extraction_rate >= 99.0


def test_extract_xlsx_named_ranges() -> None:
    """Test named range extraction (Gate G3 requirement)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"
        create_test_xlsx(file_path)

        wb_ir, stats = extract_workbook(file_path)

        # Check named ranges
        assert len(wb_ir.named_ranges) >= 1
        assert stats.named_ranges_count >= 1

        test_range = next((nr for nr in wb_ir.named_ranges if nr.name == "TestRange"), None)
        assert test_range is not None
        assert "Sheet1!$A$1:$A$2" in test_range.refers_to


def test_extract_xlsx_with_table() -> None:
    """Test table extraction."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"
        create_test_xlsx_with_table(file_path)

        wb_ir, stats = extract_workbook(file_path)

        # Check table metadata
        sheet1 = wb_ir.get_sheet_by_name("Sheet1")
        assert sheet1 is not None
        assert len(sheet1.tables) == 1

        table = sheet1.tables[0]
        assert table.name == "TestTable"
        assert table.header_row is True
        assert len(table.columns) == 3


def test_extract_xlsx_cell_types() -> None:
    """Test different cell type extraction."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        # Different cell types
        ws["A1"] = 42  # Number
        ws["A2"] = "Text"  # String
        ws["A3"] = True  # Boolean
        ws["A4"] = "=1/0"  # Error (will error)
        ws["A5"] = "=A1*2"  # Formula

        wb.save(file_path)
        wb.close()

        wb_ir, stats = extract_workbook(file_path)

        sheet = wb_ir.sheets[0]
        cells_by_type = {c.address: c.cell_type for c in sheet.cells}

        assert cells_by_type["A1"] == CellType.NUMBER
        assert cells_by_type["A2"] == CellType.STRING
        assert cells_by_type["A3"] == CellType.BOOLEAN
        assert cells_by_type["A5"] == CellType.FORMULA


def test_extract_xlsx_json_serialization() -> None:
    """Test IR JSON serialization (Phase 1.2 requirement)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"
        create_test_xlsx(file_path)

        wb_ir, stats = extract_workbook(file_path)

        # Should serialize to dict/JSON
        wb_dict = wb_ir.model_dump()
        assert isinstance(wb_dict, dict)
        assert "sheets" in wb_dict
        assert "named_ranges" in wb_dict

        # Should serialize stats
        stats_dict = stats.model_dump()
        assert isinstance(stats_dict, dict)
        assert "total_cells" in stats_dict
        # Computed properties are not serialized by default
        assert stats.formula_extraction_rate >= 0.0


def test_extract_nonexistent_file() -> None:
    """Test error handling for nonexistent files."""
    with pytest.raises(FileNotFoundError):
        extract_workbook("/nonexistent/file.xlsx")


def test_extract_unsupported_format() -> None:
    """Test error handling for unsupported formats."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.txt"
        file_path.write_text("not an excel file")

        with pytest.raises(ExtractionError):
            extract_workbook(file_path)


def test_ir_model_properties() -> None:
    """Test IR model computed properties."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"
        create_test_xlsx(file_path)

        wb_ir, stats = extract_workbook(file_path)

        # Test WorkbookIR properties
        assert wb_ir.total_cells > 0
        assert wb_ir.total_formulas >= 0
        assert wb_ir.sheet_count == 2

        # Test SheetIR properties
        sheet1 = wb_ir.sheets[0]
        assert sheet1.cell_count > 0
        assert sheet1.formula_count >= 0

        # Test ExtractionStats properties
        assert stats.formula_extraction_rate >= 0.0
        assert stats.formula_extraction_rate <= 100.0


def test_comprehensive_formula_extraction() -> None:
    """Comprehensive test for 99% formula extraction rate (Gate G3)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        # Create 100 formula cells
        for i in range(1, 101):
            row = i
            ws[f"A{row}"] = i
            ws[f"B{row}"] = f"=A{row}*2"

        wb.save(file_path)
        wb.close()

        wb_ir, stats = extract_workbook(file_path)

        # Check extraction rate
        assert stats.total_formulas == 100
        assert stats.formulas_extracted >= 99  # ≥99% = 99/100

        # Verify rate calculation
        assert stats.formula_extraction_rate >= 99.0


def test_extraction_performance() -> None:
    """Test extraction performance (memory/time in reasonable range)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = Path(tmpdir) / "test.xlsx"
        create_test_xlsx(file_path)

        wb_ir, stats = extract_workbook(file_path)

        # Should complete quickly for small files
        assert stats.extraction_time_seconds < 5.0

        # Should track stats
        assert stats.total_cells > 0
