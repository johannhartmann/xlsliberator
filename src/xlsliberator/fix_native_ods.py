"""Docker-contained post-processing for known native ODS conversion defects."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

from xlsliberator.formula_rules import FormulaRuleRegistry
from xlsliberator.lo_worker_client import LibreOfficeWorkerClient, worker_unavailable_message


class NativeODSRepairError(RuntimeError):
    """Raised when Docker-contained post-processing cannot complete truthfully."""


def _source_inventory(
    excel_path: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[str]]:
    workbook = openpyxl.load_workbook(excel_path, data_only=False)
    try:
        named_ranges: list[dict[str, str]] = []
        for name, definition in workbook.defined_names.items():
            destinations = list(definition.destinations)
            if destinations:
                sheet_name, cell_range = destinations[0]
                named_ranges.append({"name": name, "content": f"${sheet_name}.{cell_range}"})
        formula_cells: list[dict[str, str]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if (
                        isinstance(value, str)
                        and value.startswith("=")
                        and "INDIRECT" in value.upper()
                        and "ADDRESS" in value.upper()
                    ):
                        formula_cells.append({"sheet": sheet.title, "address": cell.coordinate})
        return named_ranges, formula_cells, list(workbook.sheetnames)
    finally:
        workbook.close()


def _require_worker(response: Any, operation: str) -> dict[str, Any]:
    if not response.success:
        raise NativeODSRepairError(f"{operation} failed: {worker_unavailable_message(response)}")
    return dict(response.data)


def post_process_native_ods(excel_path: Path, ods_path: Path) -> dict[str, int]:
    """Inspect and repair an ODS using only disposable Docker office workers."""
    named_ranges, candidates, excel_sheet_names = _source_inventory(excel_path)
    stats = {
        "named_ranges_added": 0,
        "formulas_scanned": len(candidates),
        "formulas_needing_fix": 0,
        "formulas_fixed": 0,
        "formulas_failed": 0,
    }
    if not named_ranges and not candidates:
        return stats

    client = LibreOfficeWorkerClient(timeout_seconds=60)
    inspection = _require_worker(
        client.request(
            {
                "op": "inspect_document_cells",
                "ods_path": str(ods_path),
                "cells": candidates,
                "timeout_seconds": 60,
            }
        ),
        "ODS formula inspection",
    )
    ods_sheet_names = [str(name) for name in inspection.get("sheet_names") or []]
    sheet_mapping = {
        excel_name: _quote_calc_sheet(ods_name)
        for excel_name, ods_name in zip(excel_sheet_names, ods_sheet_names, strict=False)
    }
    registry = FormulaRuleRegistry.with_default_rules(sheet_mapping=sheet_mapping)
    formula_repairs: list[dict[str, str]] = []
    for item in inspection.get("cells") or []:
        if not item.get("found") or int(item.get("error") or 0) != 525:
            continue
        stats["formulas_needing_fix"] += 1
        repair = registry.apply_first(str(item.get("formula") or ""))
        if repair is None or not repair.success:
            stats["formulas_failed"] += 1
            continue
        formula_repairs.append(
            {
                "sheet": str(item["sheet"]),
                "address": str(item["address"]),
                "formula": repair.after,
            }
        )

    descriptor, raw_temp = tempfile.mkstemp(
        prefix=f".{ods_path.name}.", suffix=".ods", dir=ods_path.parent
    )
    os.close(descriptor)
    repaired_path = Path(raw_temp)
    repaired_path.unlink()
    try:
        applied = _require_worker(
            client.request(
                {
                    "op": "apply_document_repairs",
                    "ods_path": str(ods_path),
                    "output_path": str(repaired_path),
                    "named_ranges": named_ranges,
                    "formula_repairs": formula_repairs,
                    "timeout_seconds": 60,
                }
            ),
            "ODS repair application",
        )
        stats["named_ranges_added"] = int(applied.get("named_ranges_added", 0))
        stats["formulas_fixed"] = int(applied.get("formulas_applied", 0))
        os.replace(repaired_path, ods_path)
    finally:
        repaired_path.unlink(missing_ok=True)
    logger.info(f"Docker-contained ODS post-processing complete: {stats}")
    return stats


def _quote_calc_sheet(sheet_name: str) -> str:
    needs_quoting = (
        not sheet_name
        or sheet_name[0].isdigit()
        or any(character in sheet_name for character in " !@#$%^&*()-+=[]{};:,.<>?/\\|`~")
    )
    return f"'{sheet_name}'" if needs_quoting else sheet_name
