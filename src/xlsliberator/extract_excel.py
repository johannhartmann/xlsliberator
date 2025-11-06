"""Excel file extraction module for various formats (.xlsx/.xlsm/.xlsb/.xls)."""

import time
from pathlib import Path
from typing import Any

import openpyxl
import pyxlsb
from loguru import logger
from openpyxl.utils import get_column_letter

from xlsliberator.ir_models import (
    CellIR,
    CellType,
    ChartMetadataIR,
    ExtractionStats,
    NamedRangeIR,
    SheetIR,
    TableMetadataIR,
    WorkbookIR,
)


class ExtractionError(Exception):
    """Raised when extraction fails."""


def extract_workbook(file_path: str | Path) -> tuple[WorkbookIR, ExtractionStats]:
    """Extract Excel workbook to intermediate representation.

    Args:
        file_path: Path to Excel file (.xlsx, .xlsm, .xlsb, .xls)

    Returns:
        Tuple of (WorkbookIR, ExtractionStats)

    Raises:
        ExtractionError: If extraction fails
        FileNotFoundError: If file doesn't exist
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = file_path.suffix.lower()
    logger.info(f"Extracting {suffix} file: {file_path}")

    start_time = time.time()

    try:
        if suffix in [".xlsx", ".xlsm"]:
            wb_ir, stats = _extract_xlsx(file_path)
        elif suffix == ".xlsb":
            wb_ir, stats = _extract_xlsb(file_path)
        elif suffix == ".xls":
            wb_ir, stats = _extract_xls(file_path)
        else:
            raise ExtractionError(f"Unsupported file format: {suffix}")

        stats.extraction_time_seconds = time.time() - start_time
        logger.success(
            f"Extracted {stats.total_cells} cells, {stats.total_formulas} formulas "
            f"in {stats.extraction_time_seconds:.2f}s"
        )

        return wb_ir, stats

    except Exception as e:
        raise ExtractionError(f"Failed to extract {file_path}: {e}") from e


def _extract_xlsx(file_path: Path) -> tuple[WorkbookIR, ExtractionStats]:
    """Extract .xlsx or .xlsm file using openpyxl.

    Args:
        file_path: Path to .xlsx/.xlsm file

    Returns:
        Tuple of (WorkbookIR, ExtractionStats)
    """
    logger.debug(f"Opening workbook with openpyxl: {file_path}")

    # Open in read-only mode for performance
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)

    # Determine if file has macros
    has_macros = file_path.suffix.lower() == ".xlsm"
    if has_macros:
        # Check for vbaProject.bin in the archive
        try:
            import zipfile

            with zipfile.ZipFile(file_path) as zf:
                has_macros = "xl/vbaProject.bin" in zf.namelist()
        except Exception:
            pass

    wb_ir = WorkbookIR(
        file_path=str(file_path),
        file_format=file_path.suffix[1:].lower(),
        has_macros=has_macros,
    )

    stats = ExtractionStats()

    # Extract named ranges
    for name, defn in wb.defined_names.items():
        if defn.value:
            named_range = NamedRangeIR(
                name=name,
                refers_to=defn.value,
                scope=None,  # Workbook-level
                comment=defn.comment if hasattr(defn, "comment") else None,
            )
            wb_ir.named_ranges.append(named_range)
            stats.named_ranges_count += 1

    # Extract sheets
    for sheet_index, ws in enumerate(wb.worksheets):
        sheet_ir = _extract_xlsx_sheet(ws, sheet_index)
        wb_ir.sheets.append(sheet_ir)

        stats.total_cells += sheet_ir.cell_count
        stats.total_formulas += sheet_ir.formula_count
        stats.formulas_extracted += sheet_ir.formula_count
        stats.tables_count += len(sheet_ir.tables)
        stats.charts_count += len(sheet_ir.charts)

    wb.close()

    return wb_ir, stats


def _extract_xlsx_sheet(ws: Any, sheet_index: int) -> SheetIR:
    """Extract a single worksheet from openpyxl.

    Args:
        ws: openpyxl worksheet object
        sheet_index: Sheet index (0-based)

    Returns:
        SheetIR object
    """
    sheet_ir = SheetIR(
        name=ws.title,
        index=sheet_index,
        visible=ws.sheet_state == "visible",
        max_row=ws.max_row or 0,
        max_col=ws.max_column or 0,
    )

    # Extract cells (iterate over all cells with data)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and cell.data_type == "n":
                continue  # Skip truly empty cells

            cell_ir = _extract_xlsx_cell(cell)
            if cell_ir:
                sheet_ir.cells.append(cell_ir)

    # Extract tables (ListObjects)
    for table in ws.tables.values():
        table_ir = TableMetadataIR(
            name=table.name,
            display_name=table.displayName,
            ref=table.ref,
            header_row=(table.headerRowCount or 0) > 0,
            totals_row=(table.totalsRowCount or 0) > 0,
            columns=[col.name for col in table.tableColumns] if table.tableColumns else [],
        )
        sheet_ir.tables.append(table_ir)

    # Extract chart metadata (basic)
    for chart in ws._charts:
        chart_ir = ChartMetadataIR(
            chart_id=str(id(chart)),
            chart_type=chart.__class__.__name__,
            title=chart.title if hasattr(chart, "title") else None,
        )
        sheet_ir.charts.append(chart_ir)

    return sheet_ir


def _extract_xlsx_cell(cell: Any) -> CellIR | None:
    """Extract a single cell from openpyxl.

    Args:
        cell: openpyxl cell object

    Returns:
        CellIR object or None if cell is empty
    """
    # Determine cell type and value
    if cell.data_type == "f":
        # Formula cell
        cell_type = CellType.FORMULA
        formula = cell.value if isinstance(cell.value, str) else None
        value = cell._value if hasattr(cell, "_value") else None
    elif cell.data_type == "n":
        # Number
        cell_type = CellType.NUMBER
        formula = None
        value = cell.value
    elif cell.data_type == "s":
        # String
        cell_type = CellType.STRING
        formula = None
        value = cell.value
    elif cell.data_type == "b":
        # Boolean
        cell_type = CellType.BOOLEAN
        formula = None
        value = cell.value
    elif cell.data_type == "e":
        # Error
        cell_type = CellType.ERROR
        formula = None
        value = cell.value
    else:
        # Empty or other
        if cell.value is None:
            return None
        cell_type = CellType.EMPTY
        formula = None
        value = cell.value

    return CellIR(
        row=cell.row - 1,  # Convert to 0-based
        col=cell.column - 1,  # Convert to 0-based
        address=cell.coordinate,
        cell_type=cell_type,
        value=value,
        formula=formula,
        comment=cell.comment.text if cell.comment else None,
    )


def _extract_xlsb(file_path: Path) -> tuple[WorkbookIR, ExtractionStats]:
    """Extract .xlsb file using pyxlsb.

    Args:
        file_path: Path to .xlsb file

    Returns:
        Tuple of (WorkbookIR, ExtractionStats)

    Note:
        pyxlsb has limited formula support - formulas are marked as unavailable
    """
    logger.debug(f"Opening workbook with pyxlsb: {file_path}")

    wb_ir = WorkbookIR(
        file_path=str(file_path),
        file_format="xlsb",
        has_macros=False,  # TODO: Detect macros in xlsb
    )

    stats = ExtractionStats()

    with pyxlsb.open_workbook(file_path) as wb:
        # Extract sheets
        for sheet_index, sheet_name in enumerate(wb.sheets):
            with wb.get_sheet(sheet_name) as ws:
                sheet_ir = SheetIR(
                    name=sheet_name,
                    index=sheet_index,
                    visible=True,  # pyxlsb doesn't expose visibility
                )

                # Extract cells
                for row_idx, row in enumerate(ws.rows()):
                    for col_idx, cell in enumerate(row):
                        if cell.v is None:
                            continue

                        # pyxlsb provides limited formula info
                        cell_type = (
                            CellType.NUMBER if isinstance(cell.v, (int, float)) else CellType.STRING
                        )

                        if cell.f:
                            # Has formula but may not be parseable
                            cell_type = CellType.FORMULA

                        cell_ir = CellIR(
                            row=row_idx,
                            col=col_idx,
                            address=f"{get_column_letter(col_idx + 1)}{row_idx + 1}",
                            cell_type=cell_type,
                            value=cell.v,
                            formula=cell.f if cell.f else None,
                        )
                        sheet_ir.cells.append(cell_ir)

                        if cell_type == CellType.FORMULA:
                            stats.total_formulas += 1
                            if cell.f:
                                stats.formulas_extracted += 1

                        stats.total_cells += 1

                sheet_ir.max_row = len(list(ws.rows()))
                wb_ir.sheets.append(sheet_ir)

    logger.warning(
        f"XLSB format has limited formula support - "
        f"{stats.formulas_extracted}/{stats.total_formulas} formulas extracted"
    )

    return wb_ir, stats


def _extract_xls(file_path: Path) -> tuple[WorkbookIR, ExtractionStats]:
    """Extract .xls file (legacy format).

    Args:
        file_path: Path to .xls file

    Returns:
        Tuple of (WorkbookIR, ExtractionStats)

    Note:
        Basic detection only - full parsing requires xlrd with additional setup
    """
    logger.warning(f"Legacy .xls format detected: {file_path}")

    wb_ir = WorkbookIR(
        file_path=str(file_path),
        file_format="xls",
        has_macros=False,  # TODO: Detect macros in xls
    )

    stats = ExtractionStats()

    # For now, just return a placeholder
    # Full implementation would require xlrd or other library
    logger.warning(".xls format requires additional libraries (xlrd) - returning empty workbook IR")

    return wb_ir, stats
