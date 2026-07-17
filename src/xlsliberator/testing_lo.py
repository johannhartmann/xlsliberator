"""LibreOffice Calc formula equivalence testing utilities."""

import math
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook

from xlsliberator.lo_worker_client import LibreOfficeWorkerClient, worker_unavailable_message


class FormulaComparisonResult:
    """Results from comparing Excel and Calc formula values."""

    def __init__(self) -> None:
        """Initialize comparison result."""
        self.total_cells: int = 0
        self.formula_cells: int = 0
        self.matching: int = 0
        self.mismatching: int = 0
        self.errors: int = 0
        self.tolerance: float = 1e-9
        self.mismatches: list[dict[str, Any]] = []

    @property
    def match_rate(self) -> float:
        """Calculate percentage of matching formula values."""
        if self.formula_cells == 0:
            return 0.0
        return (self.matching / self.formula_cells) * 100

    def summary(self) -> str:
        """Generate summary report."""
        return f"""Formula Equivalence Test Results:
- Total cells: {self.total_cells:,}
- Formula cells: {self.formula_cells:,}
- Matching: {self.matching:,} ({self.match_rate:.2f}%)
- Mismatching: {self.mismatching:,}
- Errors: {self.errors:,}
- Tolerance: {self.tolerance}
"""


def values_equal(val1: Any, val2: Any, tolerance: float = 1e-9) -> bool:
    """Compare two cell values with tolerance for floating point numbers.

    Args:
        val1: First value (from Excel)
        val2: Second value (from Calc)
        tolerance: Absolute tolerance for numeric comparisons

    Returns:
        True if values are considered equal
    """

    # Empty, zero, and empty-string results are distinct semantic outcomes.
    if val1 is None and val2 is None:
        return True
    if val1 is None or val2 is None:
        return False

    # Python bool is an int subclass; reject cross-kind comparison before numbers.
    if isinstance(val1, bool) or isinstance(val2, bool):
        return isinstance(val1, bool) and isinstance(val2, bool) and val1 == val2

    # Handle numeric values
    if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
        # Check for NaN
        if math.isnan(val1) and math.isnan(val2):
            return True
        if math.isnan(val1) or math.isnan(val2):
            return False
        # Check for infinity
        if math.isinf(val1) and math.isinf(val2):
            return val1 == val2
        if math.isinf(val1) or math.isinf(val2):
            return False
        # Compare with tolerance
        return abs(val1 - val2) <= tolerance

    # Handle string values
    if isinstance(val1, str) and isinstance(val2, str):
        return val1 == val2

    # Type mismatch
    return False


def compare_excel_calc(
    excel_path: Path,
    ods_path: Path,
    tolerance: float = 1e-9,
) -> FormulaComparisonResult:
    """Compare formula values between Excel and LibreOffice Calc files.

    Args:
        excel_path: Path to original Excel file
        ods_path: Path to converted ODS file
        tolerance: Absolute tolerance for numeric comparisons

    Returns:
        FormulaComparisonResult with comparison statistics
    """
    result = FormulaComparisonResult()
    result.tolerance = tolerance

    logger.info(f"Comparing {excel_path} with {ods_path}")

    # Load Excel workbook - need to load twice to get both formulas and values
    logger.debug(f"Loading Excel workbook: {excel_path}")
    wb_excel_formulas = load_workbook(excel_path, data_only=False)  # For formulas
    wb_excel_values = load_workbook(excel_path, data_only=True)  # For calculated values

    requested: list[dict[str, str]] = []
    source_cells: dict[tuple[str, str], tuple[Any, Any]] = {}
    try:
        for sheet_name in wb_excel_formulas.sheetnames:
            formula_sheet = wb_excel_formulas[sheet_name]
            value_sheet = wb_excel_values[sheet_name]
            for row in formula_sheet.iter_rows():
                for cell in row:
                    result.total_cells += 1
                    if cell.data_type != "f":
                        continue
                    result.formula_cells += 1
                    requested.append({"sheet": sheet_name, "address": cell.coordinate})
                    row_index = cell.row
                    column_index = cell.column
                    if not isinstance(row_index, int) or not isinstance(column_index, int):
                        raise TypeError("OpenPyXL returned a formula cell without coordinates")
                    source_cells[(sheet_name, cell.coordinate)] = (
                        cell.value,
                        value_sheet.cell(row=row_index, column=column_index).value,
                    )
    finally:
        wb_excel_formulas.close()
        wb_excel_values.close()

    response = LibreOfficeWorkerClient(timeout_seconds=60).request(
        {
            "op": "inspect_document_cells",
            "ods_path": str(ods_path),
            "cells": requested,
            "timeout_seconds": 60,
        }
    )
    if not response.success:
        raise RuntimeError(worker_unavailable_message(response))

    observed = {
        (str(item["sheet"]), str(item["address"])): item
        for item in response.data.get("cells") or []
    }
    for key, (formula, excel_value) in source_cells.items():
        item = observed.get(key)
        calc_value = item.get("value") if item and item.get("found") else None
        if item and item.get("found") and values_equal(excel_value, calc_value, tolerance):
            result.matching += 1
            continue
        result.mismatching += 1
        if item and int(item.get("error") or 0):
            result.errors += 1
        result.mismatches.append(
            {
                "sheet": key[0],
                "cell": key[1],
                "excel_value": excel_value,
                "calc_value": calc_value,
                "formula": formula,
                "target_error": item.get("error") if item else "cell_not_found",
            }
        )

    logger.success(result.summary())
    return result
