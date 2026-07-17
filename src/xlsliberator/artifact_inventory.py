"""Canonical semantic/raw workbook inventory and loss accounting."""

from __future__ import annotations

import hashlib
import importlib
import json
import posixpath

# Used only for element types and serialization; all untrusted parsing uses defusedxml.
import xml.etree.ElementTree as ET  # nosec B405
from collections import defaultdict
from collections.abc import Iterable
from contextlib import suppress
from pathlib import Path
from typing import Any, Literal
from zipfile import BadZipFile, ZipFile

import openpyxl
from defusedxml.ElementTree import fromstring as safe_fromstring

from xlsliberator.ir_models import WorkbookIR
from xlsliberator.validation_models import (
    ArtifactCoverage,
    ArtifactDisposition,
    ArtifactDispositionKind,
    CanonicalArtifactIR,
    InventoryDiff,
    SourceRef,
    TargetRef,
    WorkbookArtifactIR,
)

_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def populate_canonical_inventory(
    inventory: WorkbookArtifactIR,
    path: Path,
    *,
    vba_modules: list[dict[str, Any]] | None = None,
) -> WorkbookArtifactIR:
    """Populate stable semantic artifacts and complete raw package/stream artifacts."""
    inventory.source_sha256 = _hash_file(path)
    collector = _ArtifactCollector(path)
    collector.add(
        "workbook_metadata",
        "workbook",
        "workbook",
        ArtifactCoverage.SEMANTIC,
        {
            "format": inventory.workbook.file_format,
            "has_macros": inventory.workbook.has_macros,
            "has_external_links": inventory.workbook.has_external_links,
            "metadata": inventory.workbook.metadata,
        },
    )
    _collect_workbook_ir(collector, inventory.workbook)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        _collect_ooxml_semantics(collector, path)
        _collect_zip_package(collector, path, package_kind="opc")
    elif suffix == ".xlsb":
        _collect_zip_package(collector, path, package_kind="xlsb_opc")
        collector.add(
            "format_coverage",
            "xlsb_semantic_coverage",
            "coverage/xlsb",
            ArtifactCoverage.UNPARSED,
            {"complete": False, "raw_record_coverage": True},
        )
    elif suffix == ".xls":
        _collect_ole_streams(collector, path)
        collector.add(
            "format_coverage",
            "xls_biff_semantic_coverage",
            "coverage/xls-biff",
            ArtifactCoverage.UNPARSED,
            {"complete": False, "raw_stream_coverage": True},
        )
    elif suffix == ".ods":
        _collect_ods_semantics(collector, path)
        _collect_zip_package(collector, path, package_kind="odf")
    else:
        collector.add(
            "raw_stream",
            "unrecognized_workbook",
            "raw/file",
            ArtifactCoverage.RAW,
            raw_path=path.name,
            raw_sha256=_hash_file(path),
            raw_size=path.stat().st_size,
            known=False,
        )
    _collect_vba_semantics(collector, vba_modules or [])
    inventory.artifacts = collector.artifacts
    inventory.metadata["canonical_inventory"] = {
        "schema_version": inventory.schema_version,
        "artifact_count": len(inventory.artifacts),
        "family_counts": _family_counts(inventory.artifacts),
        "raw_coverage_complete": suffix in {".xlsx", ".xlsm", ".xlsb", ".xls", ".ods"},
        "semantic_coverage_complete": suffix in {".xlsx", ".xlsm", ".ods"},
    }
    return inventory


def inventory_ods(
    path: Path,
    *,
    role: Literal["source", "target"] = "target",
) -> WorkbookArtifactIR:
    """Create a canonical target inventory for an ODS package."""
    workbook = WorkbookIR(file_path=str(path), file_format="ods")
    inventory = WorkbookArtifactIR(inventory_role=role, workbook=workbook)
    return populate_canonical_inventory(inventory, path)


def diff_inventories(
    source: WorkbookArtifactIR,
    target: WorkbookArtifactIR,
) -> InventoryDiff:
    """Match inventories and produce one explicit disposition per source artifact."""
    target_by_key: dict[tuple[str, str], list[CanonicalArtifactIR]] = defaultdict(list)
    target_by_raw_hash: dict[str, list[CanonicalArtifactIR]] = defaultdict(list)
    for artifact in target.artifacts:
        target_by_key[(artifact.family, artifact.locator)].append(artifact)
        if artifact.raw_sha256:
            target_by_raw_hash[artifact.raw_sha256].append(artifact)

    matched: dict[str, str] = {}
    dispositions: list[ArtifactDisposition] = []
    used_target_ids: set[str] = set()
    missing: list[str] = []
    target_workbook = next(
        (item for item in target.artifacts if item.family == "workbook_metadata"),
        None,
    )
    for artifact in source.artifacts:
        exact = target_by_key.get((artifact.family, artifact.locator), [])
        raw = target_by_raw_hash.get(artifact.raw_sha256 or "", [])
        candidate = (raw or exact)[:1]
        if candidate:
            target_artifact = candidate[0]
            matched[artifact.artifact_id] = target_artifact.artifact_id
            used_target_ids.add(target_artifact.artifact_id)
            preserved = bool(
                artifact.raw_sha256
                and target_artifact.raw_sha256
                and artifact.raw_sha256 == target_artifact.raw_sha256
            )
            dispositions.append(
                ArtifactDisposition(
                    source_artifact_id=artifact.artifact_id,
                    disposition=(
                        ArtifactDispositionKind.PRESERVED
                        if preserved
                        else ArtifactDispositionKind.TRANSLATED
                    ),
                    target_refs=[_target_ref(target, target_artifact)],
                    evidence_references=["inventory-diff:stable-locator-or-content-match"],
                )
            )
            continue
        if artifact.family in {"external_link", "query", "connection"}:
            dispositions.append(
                ArtifactDisposition(
                    source_artifact_id=artifact.artifact_id,
                    disposition=ArtifactDispositionKind.EXTERNALIZED_DEPENDENCY,
                    evidence_references=["inventory-diff:external-dependency"],
                    reason="external dependency is not embedded in the target package",
                )
            )
            continue
        if artifact.family == "package_part" and artifact.known and target_workbook is not None:
            dispositions.append(
                ArtifactDisposition(
                    source_artifact_id=artifact.artifact_id,
                    disposition=ArtifactDispositionKind.TRANSLATED,
                    target_refs=[_target_ref(target, target_workbook)],
                    evidence_references=["inventory-diff:known-format-part-translated"],
                    reason="format-specific source package part was represented in the target format",
                )
            )
            continue
        missing.append(artifact.artifact_id)
        dispositions.append(
            ArtifactDisposition(
                source_artifact_id=artifact.artifact_id,
                disposition=ArtifactDispositionKind.FAILED,
                evidence_references=["inventory-diff:missing-target-artifact"],
                reason=(
                    "unknown package part was not preserved"
                    if artifact.family == "package_part" and not artifact.known
                    else "no target artifact or supported translation evidence was found"
                ),
            )
        )

    return InventoryDiff(
        source_inventory_sha256=inventory_digest(source),
        target_inventory_sha256=inventory_digest(target),
        matched=matched,
        missing_source_artifact_ids=missing,
        added_target_artifact_ids=sorted(
            artifact.artifact_id
            for artifact in target.artifacts
            if artifact.artifact_id not in used_target_ids
        ),
        dispositions=dispositions,
    )


def disposition_coverage_errors(inventory: WorkbookArtifactIR) -> list[str]:
    """Return every fail-closed disposition coverage violation."""
    dispositions: dict[str, list[ArtifactDisposition]] = defaultdict(list)
    for disposition in inventory.dispositions:
        dispositions[disposition.source_artifact_id].append(disposition)
    errors: list[str] = []
    artifact_ids = {artifact.artifact_id for artifact in inventory.artifacts}
    for artifact_id in sorted(artifact_ids):
        items = dispositions.get(artifact_id, [])
        if len(items) != 1:
            errors.append(f"{artifact_id}: expected exactly one disposition, got {len(items)}")
            continue
        item = items[0]
        if not item.target_refs and not item.evidence_references:
            errors.append(f"{artifact_id}: disposition has no target or evidence reference")
        if item.disposition is ArtifactDispositionKind.WAIVED and (
            not item.reason or not item.evidence_references
        ):
            errors.append(f"{artifact_id}: waiver requires reason and evidence")
        if item.disposition is ArtifactDispositionKind.FAILED:
            errors.append(f"{artifact_id}: disposition failed: {item.reason or 'unspecified'}")
    unknown = sorted(set(dispositions) - artifact_ids)
    errors.extend(
        f"{artifact_id}: disposition references unknown artifact" for artifact_id in unknown
    )
    return errors


def inventory_digest(inventory: WorkbookArtifactIR) -> str:
    payload = json.dumps(inventory.model_dump(mode="json"), sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode()).hexdigest()


class _ArtifactCollector:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.artifacts: list[CanonicalArtifactIR] = []
        self._ids: set[str] = set()

    def add(
        self,
        family: str,
        artifact_type: str,
        locator: str,
        coverage: ArtifactCoverage,
        semantic_data: dict[str, Any] | None = None,
        *,
        parent_artifact_id: str | None = None,
        raw_path: str | None = None,
        raw_sha256: str | None = None,
        raw_size: int | None = None,
        relationship_ids: list[str] | None = None,
        known: bool = True,
    ) -> CanonicalArtifactIR:
        artifact_id = _artifact_id(family, artifact_type, locator)
        if artifact_id in self._ids:
            raise ValueError(
                f"duplicate canonical artifact locator: {family}/{artifact_type}/{locator}"
            )
        self._ids.add(artifact_id)
        artifact = CanonicalArtifactIR(
            artifact_id=artifact_id,
            family=family,
            artifact_type=artifact_type,
            locator=locator,
            coverage=coverage,
            source_ref=SourceRef(
                source_file=str(self.path),
                artifact_type=artifact_type,
                artifact_id=artifact_id,
            ),
            parent_artifact_id=parent_artifact_id,
            semantic_data=semantic_data or {},
            raw_path=raw_path,
            raw_sha256=raw_sha256,
            raw_size=raw_size,
            relationship_ids=relationship_ids or [],
            known=known,
        )
        self.artifacts.append(artifact)
        return artifact


def _collect_workbook_ir(collector: _ArtifactCollector, workbook: WorkbookIR) -> None:
    for sheet in workbook.sheets:
        collector.add(
            "worksheet_metadata",
            "worksheet",
            f"sheet/{sheet.name}",
            ArtifactCoverage.SEMANTIC,
            {
                "name": sheet.name,
                "index": sheet.index,
                "visible": sheet.visible,
                "max_row": sheet.max_row,
                "max_col": sheet.max_col,
            },
        )
        for cell in sheet.cells:
            collector.add(
                "cell",
                "cell",
                f"sheet/{sheet.name}/cell/{cell.address}",
                ArtifactCoverage.SEMANTIC,
                cell.model_dump(mode="json"),
            )
        for table in sheet.tables:
            collector.add(
                "table",
                "table",
                f"sheet/{sheet.name}/table/{table.name}",
                ArtifactCoverage.SEMANTIC,
                table.model_dump(mode="json"),
            )
            for column in table.columns:
                collector.add(
                    "structured_reference",
                    "table_column",
                    f"sheet/{sheet.name}/table/{table.name}/column/{column}",
                    ArtifactCoverage.SEMANTIC,
                    {"table": table.name, "column": column},
                )
        for index, chart in enumerate(sheet.charts):
            collector.add(
                "chart",
                "chart",
                f"sheet/{sheet.name}/chart/{index}",
                ArtifactCoverage.SEMANTIC,
                chart.model_dump(mode="json"),
            )
    for named_range in workbook.named_ranges:
        scope = named_range.scope or "workbook"
        collector.add(
            "defined_name",
            "named_range",
            f"name/{scope}/{named_range.name}",
            ArtifactCoverage.SEMANTIC,
            named_range.model_dump(mode="json"),
        )
        if named_range.name in {"_xlnm.Print_Area", "_xlnm.Print_Titles"}:
            collector.add(
                "print_area",
                named_range.name.removeprefix("_xlnm.").lower(),
                f"print/{scope}/{named_range.name}",
                ArtifactCoverage.SEMANTIC,
                named_range.model_dump(mode="json"),
            )


def _collect_ooxml_semantics(collector: _ArtifactCollector, path: Path) -> None:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=True)
    cached = openpyxl.load_workbook(path, read_only=False, data_only=True, keep_links=True)
    try:
        properties = workbook.properties
        collector.add(
            "workbook_metadata",
            "document_properties",
            "workbook/properties",
            ArtifactCoverage.SEMANTIC,
            {
                "title": properties.title,
                "subject": properties.subject,
                "creator": properties.creator,
                "description": properties.description,
                "keywords": properties.keywords,
                "category": properties.category,
            },
        )
        calculation = workbook.calculation
        collector.add(
            "calculation_settings",
            "calculation_properties",
            "workbook/calculation",
            ArtifactCoverage.SEMANTIC,
            _public_attributes(calculation),
        )
        for sheet in workbook.worksheets:
            cached_sheet = cached[sheet.title]
            if sheet.print_area:
                collector.add(
                    "print_area",
                    "print_area",
                    f"sheet/{sheet.title}/print-area",
                    ArtifactCoverage.SEMANTIC,
                    {"range": str(sheet.print_area)},
                )
            if sheet.print_title_rows or sheet.print_title_cols:
                collector.add(
                    "print_area",
                    "print_titles",
                    f"sheet/{sheet.title}/print-titles",
                    ArtifactCoverage.SEMANTIC,
                    {
                        "rows": sheet.print_title_rows,
                        "columns": sheet.print_title_cols,
                    },
                )
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None and not cell.has_style and cell.comment is None:
                        continue
                    locator = f"sheet/{sheet.title}/cell/{cell.coordinate}"
                    collector.add(
                        "cell_semantics",
                        "cell_value_formula_cache",
                        locator,
                        ArtifactCoverage.SEMANTIC,
                        {
                            "value": _json_value(cell.value),
                            "cached_value": _json_value(cached_sheet[cell.coordinate].value),
                            "data_type": cell.data_type,
                            "number_format": cell.number_format,
                            "style_id": cell.style_id,
                            "comment": cell.comment.text if cell.comment else None,
                        },
                    )
                    if cell.has_style:
                        collector.add(
                            "style",
                            "cell_style_assignment",
                            f"{locator}/style",
                            ArtifactCoverage.SEMANTIC,
                            {"style_id": cell.style_id, "number_format": cell.number_format},
                        )
                    if cell.comment:
                        collector.add(
                            "comment",
                            "cell_comment",
                            f"{locator}/comment",
                            ArtifactCoverage.SEMANTIC,
                            {"text": cell.comment.text, "author": cell.comment.author},
                        )
                    if cell.hyperlink:
                        collector.add(
                            "hyperlink",
                            "cell_hyperlink",
                            f"{locator}/hyperlink",
                            ArtifactCoverage.SEMANTIC,
                            {
                                "target": cell.hyperlink.target,
                                "location": cell.hyperlink.location,
                                "tooltip": cell.hyperlink.tooltip,
                            },
                        )
            for row_index, row_dimension in sorted(sheet.row_dimensions.items()):
                collector.add(
                    "row_dimension",
                    "row_dimension",
                    f"sheet/{sheet.title}/row/{row_index}",
                    ArtifactCoverage.SEMANTIC,
                    _public_attributes(row_dimension),
                )
            for column_key, column_dimension in sorted(sheet.column_dimensions.items()):
                collector.add(
                    "column_dimension",
                    "column_dimension",
                    f"sheet/{sheet.title}/column/{column_key}",
                    ArtifactCoverage.SEMANTIC,
                    _public_attributes(column_dimension),
                )
            for merged in sorted(str(item) for item in sheet.merged_cells.ranges):
                collector.add(
                    "merged_cell",
                    "merged_range",
                    f"sheet/{sheet.title}/merge/{merged}",
                    ArtifactCoverage.SEMANTIC,
                    {"range": merged},
                )
            for index, rule in enumerate(sheet.conditional_formatting):
                collector.add(
                    "conditional_formatting",
                    "conditional_format",
                    f"sheet/{sheet.title}/conditional-format/{index}",
                    ArtifactCoverage.SEMANTIC,
                    {
                        "range": str(rule.sqref),
                        "rule_count": len(sheet.conditional_formatting[rule]),
                    },
                )
            for index, validation in enumerate(sheet.data_validations.dataValidation):
                collector.add(
                    "data_validation",
                    "data_validation",
                    f"sheet/{sheet.title}/validation/{index}",
                    ArtifactCoverage.SEMANTIC,
                    _public_attributes(validation),
                )
            collector.add(
                "page_setup",
                "page_setup",
                f"sheet/{sheet.title}/page-setup",
                ArtifactCoverage.SEMANTIC,
                _public_attributes(sheet.page_setup),
            )
            collector.add(
                "header_footer",
                "header_footer",
                f"sheet/{sheet.title}/header-footer",
                ArtifactCoverage.SEMANTIC,
                {
                    "odd_header": str(sheet.oddHeader),
                    "odd_footer": str(sheet.oddFooter),
                    "even_header": str(sheet.evenHeader),
                    "even_footer": str(sheet.evenFooter),
                },
            )
            for index, image in enumerate(getattr(sheet, "_images", [])):
                collector.add(
                    "image",
                    "worksheet_image",
                    f"sheet/{sheet.title}/image/{index}",
                    ArtifactCoverage.SEMANTIC,
                    {"width": image.width, "height": image.height, "format": image.format},
                )
        for index, link in enumerate(getattr(workbook, "_external_links", [])):
            collector.add(
                "external_link",
                "external_workbook_link",
                f"workbook/external-link/{index}",
                ArtifactCoverage.SEMANTIC,
                _public_attributes(link),
            )
    finally:
        cached.close()
        workbook.close()


def _collect_zip_package(
    collector: _ArtifactCollector,
    path: Path,
    *,
    package_kind: str,
) -> None:
    try:
        archive = ZipFile(path)
    except BadZipFile:
        collector.add(
            "raw_stream",
            "invalid_zip_package",
            "raw/file",
            ArtifactCoverage.RAW,
            raw_path=path.name,
            raw_sha256=_hash_file(path),
            raw_size=path.stat().st_size,
            known=False,
        )
        return
    with archive:
        names = {item.filename for item in archive.infolist() if not item.is_dir()}
        content_types = _content_types(archive) if "[Content_Types].xml" in names else {}
        relation_ids_by_source: dict[str, list[str]] = defaultdict(list)
        for name in sorted(names):
            if not name.endswith(".rels"):
                continue
            try:
                root = safe_fromstring(archive.read(name))
            except ET.ParseError:
                continue
            source = _relationship_source(name)
            for element in root.findall(f"{{{_REL_NS}}}Relationship"):
                relation_id = str(element.attrib.get("Id") or "")
                locator = f"relationship/{source}/{relation_id}"
                artifact = collector.add(
                    "relationship",
                    "package_relationship",
                    locator,
                    ArtifactCoverage.SEMANTIC_AND_RAW,
                    {
                        "source": source,
                        "id": relation_id,
                        "type": element.attrib.get("Type"),
                        "target": element.attrib.get("Target"),
                        "target_mode": element.attrib.get("TargetMode"),
                    },
                    raw_path=name,
                    raw_sha256=hashlib.sha256(ET.tostring(element)).hexdigest(),
                    raw_size=len(ET.tostring(element)),
                )
                relation_ids_by_source[source].append(artifact.artifact_id)
        for name in sorted(names):
            payload = archive.read(name)
            classification = _classify_package_part(name, package_kind)
            collector.add(
                "package_part",
                classification,
                f"package/{name}",
                ArtifactCoverage.RAW,
                {
                    "package_kind": package_kind,
                    "content_type": content_types.get(name),
                    "classification": classification,
                },
                raw_path=name,
                raw_sha256=hashlib.sha256(payload).hexdigest(),
                raw_size=len(payload),
                relationship_ids=relation_ids_by_source.get(name, []),
                known=classification != "unknown_extension",
            )
            _collect_part_family_marker(collector, name, classification, payload)


def _collect_part_family_marker(
    collector: _ArtifactCollector,
    name: str,
    classification: str,
    payload: bytes,
) -> None:
    family_map = {
        "drawing": "drawing",
        "image": "image",
        "shape": "shape",
        "text_box": "text_box",
        "pivot_table": "pivot_table",
        "pivot_cache": "pivot_cache",
        "slicer_timeline": "slicer_timeline",
        "activex": "activex",
        "control": "control",
        "vba_project": "vba_project",
        "query": "query",
        "connection": "connection",
        "data_model": "data_model",
        "embedded_ole": "embedded_ole",
        "chart": "chart_part",
    }
    family = family_map.get(classification)
    if family is None:
        return
    collector.add(
        family,
        classification,
        f"part-family/{name}",
        ArtifactCoverage.RAW,
        {"part": name},
        raw_path=name,
        raw_sha256=hashlib.sha256(payload).hexdigest(),
        raw_size=len(payload),
    )
    if classification == "drawing":
        with suppress(ET.ParseError):
            root = safe_fromstring(payload)
            for index, element in enumerate(root.iter()):
                local_name = element.tag.rsplit("}", 1)[-1]
                if local_name in {"sp", "cxnSp", "graphicFrame"}:
                    collector.add(
                        "shape",
                        "drawing_shape",
                        f"part-family/{name}/shape/{index}",
                        ArtifactCoverage.RAW,
                        {"part": name, "tag": local_name},
                        raw_path=name,
                        raw_sha256=hashlib.sha256(ET.tostring(element)).hexdigest(),
                        raw_size=len(ET.tostring(element)),
                    )
                elif local_name in {"txBody", "rich"}:
                    collector.add(
                        "text_box",
                        "drawing_text",
                        f"part-family/{name}/text/{index}",
                        ArtifactCoverage.RAW,
                        {"part": name, "tag": local_name},
                        raw_path=name,
                        raw_sha256=hashlib.sha256(ET.tostring(element)).hexdigest(),
                        raw_size=len(ET.tostring(element)),
                    )
    if classification == "vba_project":
        for family_name in ("vba_project_metadata", "vba_reference"):
            collector.add(
                family_name,
                f"raw_{family_name}",
                f"part-family/{name}/{family_name}",
                ArtifactCoverage.RAW,
                {"part": name, "semantic_parse_complete": False},
                raw_path=name,
                raw_sha256=hashlib.sha256(payload).hexdigest(),
                raw_size=len(payload),
            )
    if classification == "control" and b"macro" in payload.lower():
        collector.add(
            "event_binding",
            "raw_control_event_binding",
            f"part-family/{name}/event-binding",
            ArtifactCoverage.RAW,
            {"part": name, "semantic_parse_complete": False},
            raw_path=name,
            raw_sha256=hashlib.sha256(payload).hexdigest(),
            raw_size=len(payload),
        )


def _collect_ole_streams(collector: _ArtifactCollector, path: Path) -> None:
    try:
        olefile = importlib.import_module("olefile")
        if not olefile.isOleFile(str(path)):
            raise ValueError("not an OLE compound document")
        with olefile.OleFileIO(str(path)) as ole:
            for components in sorted(ole.listdir()):
                stream = "/".join(components)
                payload = ole.openstream(components).read()
                collector.add(
                    "raw_stream",
                    "ole_stream",
                    f"ole/{stream}",
                    ArtifactCoverage.RAW,
                    {"stream": stream},
                    raw_path=stream,
                    raw_sha256=hashlib.sha256(payload).hexdigest(),
                    raw_size=len(payload),
                    known=stream.lower() in {"workbook", "book"} or "vba" in stream.lower(),
                )
    except Exception as exc:
        collector.add(
            "raw_stream",
            "ole_container_unparsed",
            "ole/container",
            ArtifactCoverage.RAW,
            {"inspection_error": str(exc)},
            raw_path=path.name,
            raw_sha256=_hash_file(path),
            raw_size=path.stat().st_size,
            known=False,
        )


def _collect_ods_semantics(collector: _ArtifactCollector, path: Path) -> None:
    with ZipFile(path) as archive:
        root = safe_fromstring(archive.read("content.xml"))
    tag_family = {
        "table": "worksheet_metadata",
        "named-range": "defined_name",
        "named-expression": "defined_name",
        "content-validation": "data_validation",
        "conditional-format": "conditional_formatting",
        "a": "hyperlink",
        "frame": "drawing",
        "image": "image",
        "object": "embedded_ole",
        "button": "control",
        "checkbox": "control",
        "event-listener": "event_binding",
        "annotation": "comment",
    }
    counts: dict[str, int] = defaultdict(int)
    for table in (item for item in root.iter() if item.tag.rsplit("}", 1)[-1] == "table"):
        sheet_name = _attribute_by_local_name(table, "name") or f"Sheet{counts['ods_sheet'] + 1}"
        counts["ods_sheet"] += 1
        row_index = 0
        for row in (child for child in table if child.tag.rsplit("}", 1)[-1] == "table-row"):
            row_repeat = _bounded_repeat(_attribute_by_local_name(row, "number-rows-repeated"))
            for _repeated_row in range(row_repeat):
                column_index = 0
                for cell in row:
                    local_name = cell.tag.rsplit("}", 1)[-1]
                    if local_name not in {"table-cell", "covered-table-cell"}:
                        continue
                    column_repeat = _bounded_repeat(
                        _attribute_by_local_name(cell, "number-columns-repeated")
                    )
                    semantic = {
                        "formula": _attribute_by_local_name(cell, "formula"),
                        "value_type": _attribute_by_local_name(cell, "value-type"),
                        "value": _attribute_by_local_name(cell, "value"),
                        "boolean_value": _attribute_by_local_name(cell, "boolean-value"),
                        "date_value": _attribute_by_local_name(cell, "date-value"),
                        "string_value": _attribute_by_local_name(cell, "string-value"),
                        "text": "".join(cell.itertext()),
                        "style_name": _attribute_by_local_name(cell, "style-name"),
                        "covered": local_name == "covered-table-cell",
                    }
                    populated = any(value not in {None, ""} for value in semantic.values())
                    for repeat in range(column_repeat):
                        if populated:
                            address = _a1(column_index + repeat, row_index)
                            locator = f"sheet/{sheet_name}/cell/{address}"
                            collector.add(
                                "cell",
                                "cell",
                                locator,
                                ArtifactCoverage.SEMANTIC,
                                semantic,
                            )
                            collector.add(
                                "cell_semantics",
                                "cell_value_formula_cache",
                                locator,
                                ArtifactCoverage.SEMANTIC,
                                semantic,
                            )
                            if semantic["style_name"]:
                                collector.add(
                                    "style",
                                    "cell_style_assignment",
                                    f"{locator}/style",
                                    ArtifactCoverage.SEMANTIC,
                                    {"style_name": semantic["style_name"]},
                                )
                    column_index += column_repeat
                row_index += 1
    for element in root.iter():
        local_name = element.tag.rsplit("}", 1)[-1]
        family = tag_family.get(local_name)
        if family is None:
            continue
        index = counts[family]
        counts[family] += 1
        name = next(
            (value for key, value in element.attrib.items() if key.rsplit("}", 1)[-1] == "name"),
            None,
        )
        locator = f"ods/{family}/{name or index}"
        if family == "worksheet_metadata" and name:
            locator = f"sheet/{name}"
        collector.add(
            family,
            f"ods_{local_name}",
            locator,
            ArtifactCoverage.SEMANTIC,
            {"tag": local_name, "attributes": dict(sorted(element.attrib.items()))},
        )


def _collect_vba_semantics(
    collector: _ArtifactCollector, modules: Iterable[dict[str, Any]]
) -> None:
    for module in modules:
        name = str(module.get("name") or "unnamed")
        module_type = str(module.get("module_type") or "module")
        family = "vba_form" if "form" in module_type.lower() else "vba_module"
        collector.add(
            family,
            module_type,
            f"vba/module/{name}",
            ArtifactCoverage.SEMANTIC,
            dict(module),
        )


def _content_types(archive: ZipFile) -> dict[str, str]:
    root = safe_fromstring(archive.read("[Content_Types].xml"))
    defaults: dict[str, str] = {}
    overrides: dict[str, str] = {}
    for child in root:
        local = child.tag.rsplit("}", 1)[-1]
        if local == "Default":
            defaults[str(child.attrib.get("Extension") or "").lower()] = str(
                child.attrib.get("ContentType") or ""
            )
        elif local == "Override":
            overrides[str(child.attrib.get("PartName") or "").lstrip("/")] = str(
                child.attrib.get("ContentType") or ""
            )
    result = {}
    for item in archive.infolist():
        if item.is_dir():
            continue
        result[item.filename] = overrides.get(
            item.filename,
            defaults.get(item.filename.rsplit(".", 1)[-1].lower(), ""),
        )
    return result


def _relationship_source(rels_path: str) -> str:
    if rels_path == "_rels/.rels":
        return "/"
    directory, filename = posixpath.split(rels_path)
    parent = posixpath.dirname(directory)
    source_name = filename.removesuffix(".rels")
    return posixpath.join(parent, source_name)


def _classify_package_part(name: str, package_kind: str) -> str:
    lower = name.lower()
    rules = (
        ("pivotcache", "pivot_cache"),
        ("pivottable", "pivot_table"),
        ("slicer", "slicer_timeline"),
        ("timeline", "slicer_timeline"),
        ("activex", "activex"),
        ("ctrlprop", "control"),
        ("control", "control"),
        ("vbaproject", "vba_project"),
        ("query", "query"),
        ("connection", "connection"),
        ("model/", "data_model"),
        ("embeddings/", "embedded_ole"),
        ("charts/", "chart"),
        ("drawings/", "drawing"),
        ("media/", "image"),
        ("comments", "comment"),
        ("externallinks/", "external_link"),
        ("styles", "style"),
        ("tables/", "table"),
        ("worksheets/", "worksheet"),
    )
    for marker, classification in rules:
        if marker in lower:
            return classification
    known_roots = {
        "[content_types].xml",
        "_rels/",
        "docprops/",
        "xl/workbook",
        "xl/sharedstrings",
        "xl/theme/",
        "mimetype",
        "meta-inf/",
        "content.xml",
        "styles.xml",
        "meta.xml",
        "settings.xml",
        "manifest.rdf",
        "object ",
        "pictures/",
        "thumbnails/",
        "configurations2/",
    }
    if any(lower == root or lower.startswith(root) for root in known_roots):
        return f"{package_kind}_standard"
    return "unknown_extension"


def _target_ref(inventory: WorkbookArtifactIR, artifact: CanonicalArtifactIR) -> TargetRef:
    return TargetRef(
        target_file=inventory.workbook.file_path,
        artifact_type=artifact.artifact_type,
        artifact_id=artifact.artifact_id,
    )


def _attribute_by_local_name(element: ET.Element, name: str) -> str | None:
    return next(
        (value for key, value in element.attrib.items() if key.rsplit("}", 1)[-1] == name),
        None,
    )


def _bounded_repeat(raw: str | None) -> int:
    repeat = int(raw or "1")
    if repeat < 1 or repeat > 100_000:
        raise ValueError(f"ODF repeated row/column count is outside safe limits: {repeat}")
    return repeat


def _a1(column: int, row: int) -> str:
    letters = ""
    value = column + 1
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return f"{letters}{row + 1}"


def _artifact_id(family: str, artifact_type: str, locator: str) -> str:
    digest = hashlib.sha256(f"{family}\0{artifact_type}\0{locator}".encode()).hexdigest()[:24]
    return f"artifact:{family}:{digest}"


def _hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _family_counts(artifacts: Iterable[CanonicalArtifactIR]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for artifact in artifacts:
        counts[artifact.family] += 1
    return dict(sorted(counts.items()))


def _public_attributes(value: Any) -> dict[str, Any]:
    result = {}
    for name in dir(value):
        if name.startswith("_") or name in {"from_tree", "to_tree"}:
            continue
        with suppress(Exception):
            item = getattr(value, name)
            if callable(item):
                continue
            result[name] = _json_value(item)
    return result


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple, set)):
        return [_json_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    return str(value)
