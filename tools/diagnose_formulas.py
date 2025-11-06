#!/usr/bin/env python3
"""Diagnostic tool to investigate formula translation issues."""

from pathlib import Path

from loguru import logger
from openpyxl import load_workbook

from xlsliberator.uno_conn import UnoCtx, open_calc


def diagnose_formula_issues(excel_path: Path, ods_path: Path, max_samples: int = 20) -> None:
    """Diagnose formula translation issues by examining mismatches."""
    logger.info(f"Diagnosing formulas in {excel_path} vs {ods_path}")

    # Load Excel workbook
    wb_formulas = load_workbook(excel_path, data_only=False)
    wb_values = load_workbook(excel_path, data_only=True)

    # Connect to LibreOffice
    with UnoCtx() as uno_ctx:
        doc = open_calc(uno_ctx, str(ods_path))
        sheets = doc.getSheets()  # type: ignore

        sample_count = 0

        # Focus on Spielplan sheet first (showed many errors)
        sheet_name = "Spielplan"
        if sheet_name not in wb_formulas.sheetnames:
            logger.error(f"Sheet {sheet_name} not found")
            return

        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]
        sheet_calc = sheets.getByName(sheet_name)  # type: ignore

        logger.info(f"\n{'=' * 80}\nExamining sheet: {sheet_name}\n{'=' * 80}")

        # Check specific cells that showed mismatches
        problem_cells = ["T2", "U2", "V2", "W2", "X2", "Y2", "Z2", "AA2", "AB2", "AC2"]

        for cell_ref in problem_cells:
            if sample_count >= max_samples:
                break

            cell = ws_formulas[cell_ref]

            # Skip non-formula cells
            if not hasattr(cell, "data_type") or cell.data_type != "f":
                continue

            # Get Excel formula and value
            excel_formula = cell.value
            value_cell = ws_values[cell_ref]
            excel_value = value_cell.value

            # Get Calc formula and value
            calc_cell = sheet_calc.getCellByPosition(cell.column - 1, cell.row - 1)  # type: ignore
            calc_formula = calc_cell.getFormula()  # type: ignore
            calc_value = calc_cell.getValue()  # type: ignore

            # Display comparison
            logger.info(f"\n--- Cell {sheet_name}!{cell_ref} ---")
            logger.info(f"Excel Formula: {excel_formula}")
            logger.info(f"Excel Value:   {excel_value}")
            logger.info(f"Calc Formula:  {calc_formula}")
            logger.info(f"Calc Value:    {calc_value}")

            if excel_value != calc_value:
                logger.warning(f"MISMATCH: Expected {excel_value}, got {calc_value}")

            sample_count += 1

        # Check if there are named ranges
        logger.info(f"\n{'=' * 80}\nChecking Named Ranges\n{'=' * 80}")
        logger.info(f"Excel defined names: {len(wb_formulas.defined_names)}")
        for name in list(wb_formulas.defined_names.definedName)[:10]:
            logger.info(f"  - {name.name}: {name.value}")

        # Check LibreOffice named ranges
        named_ranges = doc.getPropertyValue("NamedRanges")  # type: ignore
        if named_ranges:
            logger.info(f"Calc named ranges: {named_ranges.getCount()}")  # type: ignore
            for i in range(min(10, named_ranges.getCount())):  # type: ignore
                nr = named_ranges.getByIndex(i)  # type: ignore
                logger.info(f"  - {nr.getName()}: {nr.getContent()}")  # type: ignore
        else:
            logger.warning("No named ranges found in Calc document")


if __name__ == "__main__":
    test_data_dir = Path(__file__).parent.parent / "tests" / "data"
    excel_file = test_data_dir / "Bundesliga-Ergebnis-Tippspiel_V2.31_2025-26.xlsm"
    ods_file = Path("output.ods")

    diagnose_formula_issues(excel_file, ods_file)
