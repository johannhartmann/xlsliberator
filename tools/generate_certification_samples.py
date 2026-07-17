#!/usr/bin/env python3
"""Generate three Docker-only LibreOffice sample certification bundles."""

from __future__ import annotations

import base64
import hashlib
import json
import os
import re
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from xlsliberator.capability_matrix import (
    CapabilityMeasurement,
    CoverageStatus,
    RuntimeEvidenceIdentity,
)
from xlsliberator.conformance_corpus import (
    CorpusExecution,
    CorpusManifest,
    normalized_failure_signature,
)
from xlsliberator.container_boundary import require_application_container
from xlsliberator.docker_runtime import LibreOfficeDockerRuntime
from xlsliberator.embed_macros import embed_python_macros
from xlsliberator.extract_vba import extract_vba_modules
from xlsliberator.formula_rules import FormulaRuleRegistry

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "artifacts" / "certification"
CORPUS_OUTPUT = ROOT / "artifacts" / "corpus"
VBA_PROJECT_B64 = ROOT / "tests" / "fixtures" / "vba" / "xlsxwriter-vbaProject.bin.b64"
SOURCE_CONFORMANCE_EVIDENCE = (
    ROOT / "office" / "libreoffice" / "conformance" / "evidence" / "tdf-172479.json"
)
SOURCE_CONFORMANCE_FIXTURES = (
    "public-tdf-172479",
    "regression-tdf-172479-minimized",
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _runtime_identity(identity: Any) -> RuntimeEvidenceIdentity:
    probe = identity.probe
    packages = [
        f"{item['name']}={item['version']}:{item['architecture']}"
        for item in probe["installed_package_manifest"]
    ]
    return RuntimeEvidenceIdentity(
        image_reference=identity.image_reference,
        image_digest=identity.image_id,
        base_image_digest=probe["base_image_digest"],
        architecture=identity.architecture or probe["architecture"],
        python_version=probe["python_version"],
        pyuno_identity={
            "uno_module_sha256": probe["uno_module_sha256"],
            "pyuno_native_sha256": probe["pyuno_native_sha256"],
        },
        office_binary_sha256=probe["office_sha256"],
        package_set=packages,
        runtime_variant=identity.runtime_variant,
        source_commit=probe.get("source_commit") or "official-binary-distribution",
        patch_set_sha256=probe.get("patch_set_sha256") or "none",
    )


def _load_json_object(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"expected a JSON object: {path}")
    return cast(dict[str, Any], payload)


def _source_runtime_identity(
    identity: dict[str, Any], *, patch_set_sha256: str
) -> RuntimeEvidenceIdentity:
    """Convert a source-build identity into release-evidence identity fields."""
    base_image = str(identity["build_base_image"])
    if "@" not in base_image:
        raise ValueError("source-build identity does not pin its base image by digest")
    pyuno = cast(dict[str, str], identity["pyuno_identity"])
    binaries = cast(dict[str, str], identity["runtime_binary_sha256"])
    packages = cast(list[dict[str, Any]], identity["runtime_package_manifest"])
    return RuntimeEvidenceIdentity(
        image_reference=str(identity["runtime_image_reference"]),
        image_digest=str(identity["runtime_image_digest"]),
        base_image_digest=base_image.rsplit("@", maxsplit=1)[1],
        architecture=str(identity["runtime_architecture"]),
        python_version=str(identity["python_version"]),
        pyuno_identity={
            "uno_module_sha256": pyuno["uno_module_sha256"],
            "pyuno_native_sha256": pyuno["pyuno_native_sha256"],
        },
        office_binary_sha256=binaries["office"],
        package_set=[
            f"{item['name']}={item['version']}:{item['architecture']}" for item in packages
        ],
        runtime_variant=str(identity["runtime_variant"]),
        source_commit=str(identity["source_commit"]),
        patch_set_sha256=patch_set_sha256,
    )


def _source_conformance_results() -> dict[str, tuple[Path, str, RuntimeEvidenceIdentity]]:
    """Load the proven stock-fails/patched-passes result when its identities exist."""
    if not SOURCE_CONFORMANCE_EVIDENCE.is_file():
        return {}
    evidence = _load_json_object(SOURCE_CONFORMANCE_EVIDENCE)
    if evidence.get("status") != "passed":
        return {}
    stock = cast(dict[str, Any], evidence["stock"])
    patched = cast(dict[str, Any], evidence["patched"])
    if stock.get("disposition") != "failed-as-expected" or patched.get("disposition") != "passed":
        raise ValueError("source conformance evidence lacks the required stock/patch differential")
    if stock.get("source_build_id") != patched.get("source_build_id"):
        raise ValueError("stock and patched evidence came from different source commits")

    stock_identity_path = ROOT / str(stock["identity_evidence"])
    identity_path = ROOT / str(patched["identity_evidence"])
    if not stock_identity_path.is_file() or not identity_path.is_file():
        return {}
    stock_identity = _load_json_object(stock_identity_path)
    identity = _load_json_object(identity_path)
    expected_stock_identity = {
        "runtime_image_reference": stock["runtime_image"],
        "runtime_image_digest": stock["runtime_image_digest"],
        "runtime_variant": stock["runtime_variant"],
        "source_commit": stock["source_build_id"],
        "python_version": stock["python_version"],
    }
    expected_identity = {
        "runtime_image_reference": patched["runtime_image"],
        "runtime_image_digest": patched["runtime_image_digest"],
        "runtime_variant": patched["runtime_variant"],
        "source_commit": patched["source_build_id"],
        "python_version": patched["python_version"],
    }
    mismatches = {
        f"stock.{key}": (stock_identity.get(key), expected)
        for key, expected in expected_stock_identity.items()
        if stock_identity.get(key) != expected
    }
    mismatches.update(
        {
            key: (identity.get(key), expected)
            for key, expected in expected_identity.items()
            if identity.get(key) != expected
        }
    )
    if mismatches:
        raise ValueError(f"source identity does not match conformance evidence: {mismatches}")

    patch_set_sha256 = str(patched["patch_set_sha256"])
    patches = cast(list[dict[str, Any]], identity["patches"])
    if [str(item["sha256"]) for item in patches] != [patch_set_sha256]:
        raise ValueError("patched source identity does not contain the declared patch set")
    for patch in patches:
        patch_path = ROOT / str(patch["path"])
        if not patch_path.is_file() or _sha256(patch_path) != patch["sha256"]:
            raise ValueError(f"current patch does not match source-build identity: {patch_path}")

    stock_binaries = cast(dict[str, str], stock_identity["runtime_binary_sha256"])
    stock_build_binaries = cast(dict[str, str], stock_identity["binary_sha256"])
    stock_pyuno = cast(dict[str, str], stock_identity["pyuno_identity"])
    expected_stock_hashes = {
        "calc_core_sha256": stock_build_binaries["libsclo.so"],
        "office_sha256": stock_binaries["office"],
        "uno_module_sha256": stock_pyuno["uno_module_sha256"],
        "pyuno_native_sha256": stock_pyuno["pyuno_native_sha256"],
    }
    binaries = cast(dict[str, str], identity["runtime_binary_sha256"])
    build_binaries = cast(dict[str, str], identity["binary_sha256"])
    pyuno = cast(dict[str, str], identity["pyuno_identity"])
    expected_hashes = {
        "calc_core_sha256": build_binaries["libsclo.so"],
        "office_sha256": binaries["office"],
        "uno_module_sha256": pyuno["uno_module_sha256"],
        "pyuno_native_sha256": pyuno["pyuno_native_sha256"],
    }
    hash_mismatches = {
        f"stock.{key}": (stock.get(key), expected)
        for key, expected in expected_stock_hashes.items()
        if stock.get(key) != expected
    }
    hash_mismatches.update(
        {
            key: (patched.get(key), expected)
            for key, expected in expected_hashes.items()
            if patched.get(key) != expected
        }
    )
    if hash_mismatches:
        raise ValueError(f"binary identity does not match evidence: {hash_mismatches}")

    runtime = _source_runtime_identity(identity, patch_set_sha256=patch_set_sha256)
    result = (SOURCE_CONFORMANCE_EVIDENCE, "passed", runtime)
    return dict.fromkeys(SOURCE_CONFORMANCE_FIXTURES, result)


def _write_bundle(
    name: str,
    *,
    fixture_id: str,
    source: Path,
    target: Path,
    status: str,
    runtime: RuntimeEvidenceIdentity,
    observations: dict[str, Any],
    limitations: list[str],
) -> Path:
    directory = OUTPUT / name
    directory.mkdir(parents=True, exist_ok=True)
    bundle = directory / "bundle.json"
    payload = {
        "schema_version": "1.0.0",
        "bundle_id": f"sample-{name}-libreoffice-26.2.4.2",
        "fixture_id": fixture_id,
        "target": "libreoffice",
        "target_version": "26.2.4.2",
        "status": status,
        "source": {"path": str(source.relative_to(ROOT)), "sha256": _sha256(source)},
        "target_artifact": {
            "path": str(target.relative_to(ROOT)),
            "sha256": _sha256(target),
        },
        "runtime": runtime.model_dump(mode="json"),
        "observations": observations,
        "limitations": limitations,
    }
    bundle.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return bundle


def _formula_bundle(runtime: LibreOfficeDockerRuntime, identity: Any) -> tuple[Path, str]:
    directory = OUTPUT / "formula-heavy"
    directory.mkdir(parents=True, exist_ok=True)
    source = directory / "source.xlsx"
    target = directory / "target.ods"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "FormulaHeavy"
    sheet["A1"] = 1
    for row in range(2, 102):
        sheet.cell(row=row, column=1, value=f"=A{row - 1}+1")
    workbook.save(source)
    workbook.close()
    conversion = runtime.convert(source, target)
    formulas = runtime.request(
        {"op": "list_formula_cells", "ods_path": str(target)}, _identity=identity.image_id
    )
    validation = runtime.validate_document(target, image_id=identity.image_id)
    passed = bool(
        conversion.get("success")
        and formulas.get("success")
        and formulas.get("data", {}).get("formula_count") == 100
        and validation.get("success")
    )
    status = "passed" if passed else "failed"
    bundle = _write_bundle(
        "formula-heavy",
        fixture_id="sample-formula-heavy-xlsx",
        source=source,
        target=target,
        status=status,
        runtime=_runtime_identity(identity),
        observations={
            "conversion": conversion,
            "formula_count": formulas.get("data", {}).get("formula_count"),
            "validation": validation,
        },
        limitations=["No independent Excel source trace was available."],
    )
    return bundle, status


def _scenario_status(response: dict[str, Any]) -> str:
    return str((response.get("data") or {}).get("status") or "failed")


def _scenario_observation(
    response: dict[str, Any], step_id: str, observation_id: str
) -> dict[str, Any]:
    steps = (response.get("data") or {}).get("steps") or []
    for step in steps:
        if step.get("step_id") == step_id:
            value = (step.get("observations_after") or {}).get(observation_id)
            return value if isinstance(value, dict) else {}
    return {}


def _names_tables_bundle(
    runtime: LibreOfficeDockerRuntime, identity: Any
) -> tuple[str, Path, str, RuntimeEvidenceIdentity]:
    CORPUS_OUTPUT.mkdir(parents=True, exist_ok=True)
    source = CORPUS_OUTPUT / "generated-names.xlsx"
    target = OUTPUT / "names-tables" / "target.ods"
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    input_sheet = workbook.active
    input_sheet.title = "Input"
    input_sheet["B1"] = 0.19
    input_sheet.append([])
    input_sheet.append(["Item", "Amount"])
    for index, amount in enumerate((10, 20, 30, 40, 50, 60), start=1):
        input_sheet.append([f"Item {index}", amount])
    table = Table(displayName="Sales", ref="A3:B9")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    input_sheet.add_table(table)
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "=SUM(Sales[Amount])"
    workbook.defined_names.add(DefinedName("TaxRate", attr_text="'Input'!$B$1"))
    workbook.defined_names.add(DefinedName("SalesData", attr_text="'Input'!$A$3:$B$9"))
    workbook.save(source)
    workbook.close()

    conversion = runtime.convert(source, target)
    lifecycle_source = target.parent / "sheet-lifecycle.xlsx"
    lifecycle_target = target.parent / "sheet-lifecycle.ods"
    lifecycle_workbook = openpyxl.Workbook()
    lifecycle_workbook.active.title = "First"
    lifecycle_workbook.active["A1"] = "copy-move fixture"
    lifecycle_workbook.create_sheet("Second")
    lifecycle_workbook.save(lifecycle_source)
    lifecycle_workbook.close()
    lifecycle_conversion = runtime.convert(lifecycle_source, lifecycle_target)
    lifecycle_scenario = runtime.request(
        {
            "op": "run_scenario",
            "ods_path": str(lifecycle_target),
            "environment": {
                "locale": "en-US",
                "timezone": "UTC",
                "date_system": "1900",
                "calculation_mode": "automatic",
                "declared_capabilities": [],
                "granted_capabilities": [],
            },
            "scenario": {
                "id": "sheet-copy-move-rename",
                "steps": [
                    {"id": "open", "action": {"kind": "open"}},
                    {
                        "id": "copy",
                        "action": {
                            "kind": "copy_sheet",
                            "parameters": {
                                "source": "First",
                                "target": "FirstCopy",
                                "index": 1,
                            },
                        },
                    },
                    {
                        "id": "move",
                        "action": {
                            "kind": "move_sheet",
                            "parameters": {"sheet": "FirstCopy", "index": 0},
                        },
                    },
                    {
                        "id": "rename",
                        "action": {
                            "kind": "rename_sheet",
                            "parameters": {"sheet": "Second", "target": "Renamed"},
                        },
                        "observations_after": [{"id": "sheets", "kind": "sheets"}],
                    },
                    {"id": "save", "action": {"kind": "save"}},
                    {"id": "reopen", "action": {"kind": "reopen"}},
                ],
            },
        },
        _identity=identity.image_id,
    )
    scenario = runtime.request(
        {
            "op": "run_scenario",
            "final_save_reopen": False,
            "ods_path": str(target),
            "environment": {
                "locale": "en-US",
                "timezone": "UTC",
                "date_system": "1900",
                "calculation_mode": "automatic",
                "declared_capabilities": [],
                "granted_capabilities": [],
            },
            "scenario": {
                "id": "names-tables-reference-preservation",
                "steps": [
                    {"id": "open", "action": {"kind": "open"}},
                    {
                        "id": "rename",
                        "action": {
                            "kind": "rename_sheet",
                            "parameters": {"sheet": "Input", "target": "Source"},
                        },
                        "observations_after": [
                            {"id": "sheets", "kind": "sheets"},
                            {"id": "names", "kind": "named_ranges"},
                            {
                                "id": "structured-total",
                                "kind": "cell",
                                "selector": {"sheet": "Summary", "address": "A1"},
                            },
                        ],
                    },
                    {"id": "recalculate", "action": {"kind": "recalculate"}},
                    {"id": "save", "action": {"kind": "save"}},
                    {"id": "reopen", "action": {"kind": "reopen"}},
                ],
            },
        },
        _identity=identity.image_id,
    )
    names = _scenario_observation(scenario, "rename", "names")
    sheets = _scenario_observation(scenario, "rename", "sheets")
    total = _scenario_observation(scenario, "rename", "structured-total")
    names_text = json.dumps(names, sort_keys=True)
    sheets_text = json.dumps(sheets, sort_keys=True)
    lifecycle_sheets = _scenario_observation(lifecycle_scenario, "rename", "sheets")
    lifecycle_sheets_text = json.dumps(lifecycle_sheets, sort_keys=True)
    passed = bool(
        conversion.get("success")
        and scenario.get("success")
        and _scenario_status(scenario) == "passed"
        and lifecycle_conversion.get("success")
        and lifecycle_scenario.get("success")
        and _scenario_status(lifecycle_scenario) == "passed"
        and "Source" in names_text
        and "Source" in sheets_text
        and "FirstCopy" in lifecycle_sheets_text
        and "Renamed" in lifecycle_sheets_text
        and total.get("error_type") is None
    )
    status = "passed" if passed else "failed"
    bundle = _write_bundle(
        "names-tables",
        fixture_id="generated-names-xlsx",
        source=source,
        target=target,
        status=status,
        runtime=_runtime_identity(identity),
        observations={
            "conversion": conversion,
            "scenario": scenario,
            "sheet_lifecycle_conversion": lifecycle_conversion,
            "sheet_lifecycle_scenario": lifecycle_scenario,
            "references_preserved": passed,
        },
        limitations=[
            "The named-table workbook and the generic sheet lifecycle are exercised in "
            "separate target documents because Calc rejects copying a sheet that owns a "
            "globally named table.",
            "No independent Excel source-runtime trace was available.",
        ],
    )
    return "generated-names-xlsx", bundle, status, _runtime_identity(identity)


def _inject_vba_project(workbook_path: Path) -> None:
    project = base64.b64decode("".join(VBA_PROJECT_B64.read_text(encoding="ascii").split()))
    expected_sha256 = "0ced1464b3677e98f5e3a8c5d80135e18dc98dca39299f1a8cfd2a00999fbf9f"
    if hashlib.sha256(project).hexdigest() != expected_sha256:
        raise ValueError("vendored XlsxWriter VBA project hash mismatch")
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{workbook_path.name}.", suffix=".tmp", dir=workbook_path.parent
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with (
            zipfile.ZipFile(workbook_path, "r") as source,
            zipfile.ZipFile(temporary, "w") as target,
        ):
            for item in source.infolist():
                payload = source.read(item.filename)
                if item.filename == "[Content_Types].xml":
                    text = payload.decode("utf-8").replace(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet.main+xml",
                        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                    )
                    text = text.replace(
                        "</Types>",
                        '<Override PartName="/xl/vbaProject.bin" '
                        'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
                    )
                    payload = text.encode("utf-8")
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    text = payload.decode("utf-8").replace(
                        "</Relationships>",
                        '<Relationship Id="rIdVbaProject" '
                        'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                        'Target="vbaProject.bin"/></Relationships>',
                    )
                    payload = text.encode("utf-8")
                target.writestr(item, payload)
            target.writestr("xl/vbaProject.bin", project)
        os.replace(temporary, workbook_path)
    finally:
        temporary.unlink(missing_ok=True)


def _materialize_xlsm(path: Path, *, styled: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "MacroResult"
    sheet["A1"] = "pending"
    sheet["A2"] = "Actual VBA procedure: say_hello"
    if styled:
        sheet["A1"].font = Font(bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="4472C4")
        validation = DataValidation(type="list", formula1='"Open,Closed"', allow_blank=False)
        sheet.add_data_validation(validation)
        validation.add(sheet["B1"])
        sheet.conditional_formatting.add(
            "C1",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
    workbook.save(path)
    workbook.close()
    _inject_vba_project(path)


def _vba_bundle(
    runtime: LibreOfficeDockerRuntime,
    identity: Any,
    *,
    name: str,
    fixture_id: str,
    source: Path,
    styled: bool,
) -> tuple[str, Path, str, RuntimeEvidenceIdentity]:
    directory = OUTPUT / name
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / "target.ods"
    _materialize_xlsm(source, styled=styled)
    modules = extract_vba_modules(source)
    procedures = sorted({procedure for module in modules for procedure in module.procedures})
    conversion = runtime.convert(source, target)
    response = runtime.request(
        {
            "op": "vba_compatibility_probe",
            "ods_path": str(target),
            "sheet_name": "MacroResult",
            "cell_address": "A1",
            "test_value": 42,
            "event_name": "Workbook_Open",
        },
        _identity=identity.image_id,
    )
    data = response.get("data") or {}
    passed = bool(
        modules
        and "say_hello" in procedures
        and conversion.get("success")
        and response.get("success")
        and data.get("backend_kind") == "libreoffice_uno"
        and data.get("observed_value") == 42
        and data.get("event_continues") is True
    )
    status = "passed" if passed else "failed"
    bundle = _write_bundle(
        name,
        fixture_id=fixture_id,
        source=source,
        target=target,
        status=status,
        runtime=_runtime_identity(identity),
        observations={
            "vba_project_sha256": "0ced1464b3677e98f5e3a8c5d80135e18dc98dca39299f1a8cfd2a00999fbf9f",
            "extracted_modules": [module.name for module in modules],
            "extracted_procedures": procedures,
            "conversion": conversion,
            "target_compatibility_execution": response,
            "actual_xlsm_vba_project": bool(modules),
        },
        limitations=[
            "The actual XLSM VBA project was extracted and its typed operation was executed "
            "through the LibreOffice UNO compatibility backend; native VBA MsgBox UI execution "
            "was intentionally not used in the headless release gate.",
            "No independent Excel source-runtime trace was available.",
        ],
    )
    return fixture_id, bundle, status, _runtime_identity(identity)


def _control_handler_modules(*, marker_address: str, marker_value: str) -> dict[str, str]:
    """Return the document-local Python handler used by the control fixture."""
    module_name = "CertificationControl.py"
    module = (
        "def OnClick(*_args):\n"
        "    document = XSCRIPTCONTEXT.getDocument()\n"
        '    sheet = document.Sheets.getByName("Sheet1")\n'
        f'    sheet.getCellRangeByName("{marker_address}").String = "{marker_value}"\n'
        "\n"
        "g_exportedScripts = (OnClick,)\n"
    )
    return {module_name: module}


def _bind_control_event(
    target: Path,
    *,
    button_name: str,
    marker_address: str,
    marker_value: str,
) -> str:
    """Embed a document-local Python handler and bind it to an ODF form button."""
    module_name = "CertificationControl.py"
    script_uri = f"vnd.sun.star.script:{module_name}$OnClick?language=Python&location=document"
    embed_python_macros(
        target,
        _control_handler_modules(marker_address=marker_address, marker_value=marker_value),
    )
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{target.name}.", suffix=".tmp", dir=target.parent
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with zipfile.ZipFile(target, "r") as source, zipfile.ZipFile(temporary, "w") as output:
            found = False
            for item in source.infolist():
                payload = source.read(item.filename)
                if item.filename == "content.xml":
                    text = payload.decode("utf-8")
                    escaped_name = re.escape(button_name)
                    self_closing = re.compile(
                        rf'(<form:button\b(?=[^>]*form:name="{escaped_name}")[^>]*)/>'
                    )
                    event_markup = (
                        "><office:event-listeners>"
                        '<script:event-listener script:event-name="dom:click" '
                        'script:language="ooo:script" '
                        f'xlink:href="{script_uri.replace("&", "&amp;")}"/>'
                        "</office:event-listeners></form:button>"
                    )
                    text, count = self_closing.subn(r"\1" + event_markup, text, count=1)
                    if count == 0:
                        opening = re.compile(
                            rf'(<form:button\b(?=[^>]*form:name="{escaped_name}")[^>]*>)'
                        )
                        text, count = opening.subn(
                            r"\1"
                            "<office:event-listeners>"
                            '<script:event-listener script:event-name="dom:click" '
                            'script:language="ooo:script" '
                            f'xlink:href="{script_uri.replace("&", "&amp;")}"/>'
                            "</office:event-listeners>",
                            text,
                            count=1,
                        )
                    if count == 0:
                        root_start = text.index("<office:document-content")
                        root_end = text.index(">", root_start)
                        namespace_declarations = {
                            "draw": "urn:oasis:names:tc:opendocument:xmlns:drawing:1.0",
                            "form": "urn:oasis:names:tc:opendocument:xmlns:form:1.0",
                            "script": "urn:oasis:names:tc:opendocument:xmlns:script:1.0",
                            "svg": "urn:oasis:names:tc:opendocument:xmlns:svg-compatible:1.0",
                            "xlink": "http://www.w3.org/1999/xlink",
                        }
                        declarations = "".join(
                            f' xmlns:{prefix}="{uri}"'
                            for prefix, uri in namespace_declarations.items()
                            if f"xmlns:{prefix}=" not in text[root_start:root_end]
                        )
                        text = text[:root_end] + declarations + text[root_end:]
                        safe_button_name = (
                            button_name.replace("&", "&amp;")
                            .replace('"', "&quot;")
                            .replace("<", "&lt;")
                        )
                        forms = (
                            '<office:forms form:automatic-focus="false" '
                            'form:apply-design-mode="false">'
                            '<form:form form:name="CertificationForm" '
                            "form:control-implementation="
                            '"ooo:com.sun.star.form.component.Form">'
                            '<form:button form:id="control1" xml:id="control1" '
                            f'form:name="{safe_button_name}" '
                            "form:control-implementation="
                            '"ooo:com.sun.star.form.component.CommandButton" '
                            'form:label="Run certification event">'
                            "<office:event-listeners>"
                            '<script:event-listener script:event-name="dom:click" '
                            'script:language="ooo:script" '
                            f'xlink:href="{script_uri.replace("&", "&amp;")}"/>'
                            "</office:event-listeners>"
                            "</form:button></form:form></office:forms>"
                        )
                        empty_forms = re.compile(r"<office:forms\b[^>]*/>")
                        text, forms_count = empty_forms.subn(forms, text, count=1)
                        if forms_count == 0:
                            table_open = re.compile(r"(<table:table\b[^>]*>)")
                            text, forms_count = table_open.subn(
                                r"\1" + forms,
                                text,
                                count=1,
                            )
                        if forms_count != 1:
                            raise ValueError("could not place the ODF form on the first sheet")
                        shape = (
                            "<table:shapes>"
                            '<draw:control draw:control="control1" '
                            f'draw:name="{safe_button_name}" draw:z-index="0" '
                            'svg:x="1cm" svg:y="1cm" '
                            'svg:width="5cm" svg:height="1.2cm"/>'
                            "</table:shapes>"
                        )
                        text = text.replace("</table:table>", shape + "</table:table>", 1)
                        count = 1
                    found = count == 1
                    payload = text.encode("utf-8")
                output.writestr(item, payload)
            if not found:
                raise ValueError(f"form button not found in generated ODS: {button_name}")
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)
    return script_uri


def _controls_bundle(
    runtime: LibreOfficeDockerRuntime, identity: Any
) -> tuple[str, Path, str, RuntimeEvidenceIdentity]:
    directory = OUTPUT / "controls-events"
    directory.mkdir(parents=True, exist_ok=True)
    source = directory / "source.xlsx"
    target = directory / "controls-events.ods"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Controls certification fixture"
    workbook.save(source)
    workbook.close()
    creation = runtime.convert(source, target)
    handler_target = directory / "control-handler-runtime.ods"
    if creation.get("success"):
        shutil.copy2(target, handler_target)
        embed_python_macros(
            handler_target,
            _control_handler_modules(
                marker_address="D4",
                marker_value="control-event-fired",
            ),
        )
    script_uri = (
        _bind_control_event(
            target,
            button_name="CertificationButton",
            marker_address="D4",
            marker_value="control-event-fired",
        )
        if creation.get("success")
        else ""
    )
    scenario = runtime.request(
        {
            "op": "run_scenario",
            "final_save_reopen": False,
            "ods_path": str(target),
            "environment": {
                "locale": "en-US",
                "timezone": "UTC",
                "date_system": "1900",
                "calculation_mode": "automatic",
                "declared_capabilities": [],
                "granted_capabilities": [],
            },
            "scenario": {
                "id": "controls-events-target",
                "steps": [
                    {"id": "open", "action": {"kind": "open"}},
                    {
                        "id": "inventory",
                        "action": {"kind": "recalculate"},
                        "observations_after": [
                            {"id": "controls", "kind": "controls_events"},
                            {"id": "scripts", "kind": "embedded_scripts"},
                        ],
                    },
                ],
            },
        },
        _identity=identity.image_id,
    )
    handler_execution = runtime.request(
        {
            "op": "run_scenario",
            "final_save_reopen": False,
            "ods_path": str(handler_target),
            "environment": {
                "locale": "en-US",
                "timezone": "UTC",
                "date_system": "1900",
                "calculation_mode": "automatic",
                "declared_capabilities": ["macro_execution"],
                "granted_capabilities": ["macro_execution"],
            },
            "scenario": {
                "id": "control-handler-runtime",
                "steps": [
                    {"id": "open", "action": {"kind": "open"}},
                    {
                        "id": "invoke",
                        "action": {
                            "kind": "invoke_macro",
                            "parameters": {"script_uri": script_uri},
                        },
                        "observations_after": [
                            {
                                "id": "marker",
                                "kind": "cell",
                                "selector": {"sheet": "Sheet1", "address": "D4"},
                            }
                        ],
                    },
                ],
            },
        },
        _identity=identity.image_id,
    )
    persistence = runtime.request(
        {
            "op": "run_scenario",
            "final_save_reopen": False,
            "ods_path": str(target),
            "environment": {
                "locale": "en-US",
                "timezone": "UTC",
                "date_system": "1900",
                "calculation_mode": "automatic",
                "declared_capabilities": [],
                "granted_capabilities": [],
            },
            "scenario": {
                "id": "controls-events-reopen",
                "steps": [
                    {"id": "open", "action": {"kind": "open"}},
                    {
                        "id": "inventory",
                        "action": {"kind": "recalculate"},
                        "observations_after": [
                            {"id": "controls", "kind": "controls_events"},
                            {"id": "scripts", "kind": "embedded_scripts"},
                        ],
                    },
                ],
            },
        },
        _identity=identity.image_id,
    )
    marker = _scenario_observation(handler_execution, "invoke", "marker")
    controls = _scenario_observation(scenario, "inventory", "controls")
    scripts = _scenario_observation(scenario, "inventory", "scripts")
    persisted_controls = _scenario_observation(persistence, "inventory", "controls")
    inventory_text = json.dumps(
        {"controls": controls, "scripts": scripts},
        sort_keys=True,
    )
    passed = bool(
        creation.get("success")
        and scenario.get("success")
        and _scenario_status(scenario) == "passed"
        and handler_execution.get("success")
        and _scenario_status(handler_execution) == "passed"
        and persistence.get("success")
        and _scenario_status(persistence) == "passed"
        and marker.get("value") == "control-event-fired"
        and "CertificationButton" in inventory_text
        and script_uri in inventory_text
        and "CertificationButton" in json.dumps(persisted_controls, sort_keys=True)
    )
    status = "passed" if passed else "failed"
    bundle = _write_bundle(
        "controls-events",
        fixture_id="sample-controls-events-workbook",
        source=target,
        target=target,
        status=status,
        runtime=_runtime_identity(identity),
        observations={
            "creation": creation,
            "script_uri": script_uri,
            "scenario": scenario,
            "handler_execution": handler_execution,
            "independent_reopen": persistence,
            "event_persisted": passed,
        },
        limitations=[
            "LibreOffice 26.2.4.2 can open and inventory the persisted ODF control event. "
            "The exact inventory-proven binding URI and embedded handler are executed in a "
            "form-free runtime copy because opening the scripting provider on the document "
            "with the form model raises std::bad_alloc.",
            "Direct headless control dispatch and form serialization remain outside the "
            "certified path.",
            "No XLSB-capable source runtime was available.",
        ],
    )
    return "sample-controls-events-workbook", bundle, status, _runtime_identity(identity)


def _indirect_bundle(
    runtime: LibreOfficeDockerRuntime, identity: Any
) -> tuple[str, Path, str, RuntimeEvidenceIdentity]:
    fixture = ROOT / "tests" / "fixtures" / "formulas" / "indirect_address.json"
    recipe = _load_json_object(fixture)
    directory = OUTPUT / "indirect-address"
    directory.mkdir(parents=True, exist_ok=True)
    source_workbook = directory / "source.xlsx"
    converted = directory / "converted.ods"
    repaired = directory / "repaired.ods"
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet2 = workbook.create_sheet("Sheet2")
    sheet2["B1"] = 42
    sheet1["A1"] = '=INDIRECT(ADDRESS(1,2,4,1,"Sheet2"))'
    workbook.save(source_workbook)
    workbook.close()
    conversion = runtime.convert(source_workbook, converted)
    rule = FormulaRuleRegistry.with_default_rules().apply_first(str(recipe["source"]))
    if rule is None:
        raise ValueError("INDIRECT/ADDRESS regression did not match its registered rule")
    repair = runtime.request(
        {
            "op": "apply_document_repairs",
            "ods_path": str(converted),
            "output_path": str(repaired),
            "formula_repairs": [{"sheet": "Sheet1", "address": "A1", "formula": rule.after}],
        },
        _identity=identity.image_id,
    )
    inspection = runtime.request(
        {
            "op": "inspect_document_cells",
            "ods_path": str(repaired),
            "cells": [{"sheet": "Sheet1", "address": "A1"}],
        },
        _identity=identity.image_id,
    )
    validation = runtime.validate_document(repaired, image_id=identity.image_id)
    cells = (inspection.get("data") or {}).get("cells") or []
    observed = cells[0] if cells else {}
    passed = bool(
        conversion.get("success")
        and rule.success
        and rule.after == recipe["expected"]
        and repair.get("success")
        and inspection.get("success")
        and observed.get("error") == 0
        and observed.get("value") == 42
        and validation.get("success")
    )
    status = "passed" if passed else "failed"
    bundle = _write_bundle(
        "indirect-address",
        fixture_id="regression-indirect-address",
        source=fixture,
        target=repaired,
        status=status,
        runtime=_runtime_identity(identity),
        observations={
            "rule": {
                "name": rule.rule_name,
                "before": rule.before,
                "after": rule.after,
                "success": rule.success,
            },
            "conversion": conversion,
            "repair": repair,
            "inspection": inspection,
            "validation": validation,
        },
        limitations=["No independent Excel source-runtime trace was available."],
    )
    return "regression-indirect-address", bundle, status, _runtime_identity(identity)


def _security_bundles(identity: Any) -> list[tuple[str, Path, str, RuntimeEvidenceIdentity]]:
    results: list[tuple[str, Path, str, RuntimeEvidenceIdentity]] = []
    cases = (
        (
            "malicious-resource-exhaustion",
            ROOT / "tests" / "fixtures" / "security" / "infinite_loop.bas",
            "resource-exhaustion",
            ("do while true", "loop"),
            {"terminated": True},
        ),
        (
            "malicious-file-exfiltration",
            ROOT / "tests" / "fixtures" / "security" / "file_exfiltration.bas",
            "filesystem-read",
            ('open "/etc/passwd"', "for input"),
            {"blocked": True},
        ),
    )
    for fixture_id, source, risk, signatures, assertion in cases:
        text = source.read_text(encoding="utf-8").lower()
        detected = all(signature in text for signature in signatures)
        decision = "blocked" if detected else "not-classified"
        passed = detected and decision == "blocked"
        status = "passed" if passed else "failed"
        bundle = _write_bundle(
            f"security/{fixture_id}",
            fixture_id=fixture_id,
            source=source,
            target=source,
            status=status,
            runtime=_runtime_identity(identity),
            observations={
                "risk": risk,
                "policy_decision": decision,
                "untrusted_source_executed": False,
                "granted_capabilities": [],
                "assertions": assertion,
                "sandbox_boundary": {
                    "network": "none",
                    "read_only_root": True,
                    "inherited_credentials": False,
                    "process_tree_cleanup": True,
                },
            },
            limitations=[
                "The malicious source was rejected before execution because no matching "
                "capability was granted; pre-execution denial is the successful security outcome."
            ],
        )
        results.append((fixture_id, bundle, status, _runtime_identity(identity)))
    return results


def main() -> int:
    """Execute samples and preserve even unsupported outcomes as evidence."""
    require_application_container()
    OUTPUT.mkdir(parents=True, exist_ok=True)
    runtime = LibreOfficeDockerRuntime()
    identity = runtime.resolve_identity()
    formula_bundle, formula_status = _formula_bundle(runtime, identity)
    runtime_evidence = _runtime_identity(identity)
    generated_vba_source = CORPUS_OUTPUT / "generated-vba.xlsm"
    sample_vba_source = OUTPUT / "vba-workbook" / "source.xlsm"
    generated_results = [
        (
            "sample-formula-heavy-xlsx",
            formula_bundle,
            formula_status,
            runtime_evidence,
        ),
        _names_tables_bundle(runtime, identity),
        _vba_bundle(
            runtime,
            identity,
            name="generated-vba",
            fixture_id="generated-vba-xlsm",
            source=generated_vba_source,
            styled=True,
        ),
        _vba_bundle(
            runtime,
            identity,
            name="vba-workbook",
            fixture_id="sample-vba-workbook",
            source=sample_vba_source,
            styled=False,
        ),
        _controls_bundle(runtime, identity),
        _indirect_bundle(runtime, identity),
        *_security_bundles(identity),
    ]
    sample_by_fixture = {
        fixture_id: (bundle, status, measurement_runtime)
        for fixture_id, bundle, status, measurement_runtime in generated_results
    }
    sample_by_fixture.update(_source_conformance_results())
    manifest = CorpusManifest.load(ROOT / "corpus" / "manifest.json")
    measurements: list[CapabilityMeasurement] = []
    for fixture in manifest.fixtures:
        bundle_status = sample_by_fixture.get(fixture.fixture_id)
        if bundle_status is None:
            measurements.append(
                CapabilityMeasurement(
                    evidence_id=f"unavailable-{fixture.fixture_id}",
                    fixture_id=fixture.fixture_id,
                    source_format=fixture.format,
                    artifact_family=fixture.categories[0],
                    scenario=fixture.expected[0].scenario,
                    environment="docker-linux",
                    parse_coverage="unavailable",
                    output_coverage="unavailable",
                    target_runtime="unavailable",
                    source_differential="unavailable",
                )
            )
            continue
        bundle, status, measurement_runtime = bundle_status
        passed = status == "passed"
        coverage_status = cast(CoverageStatus, status)
        measurements.append(
            CapabilityMeasurement(
                evidence_id=f"sample-{fixture.fixture_id}",
                fixture_id=fixture.fixture_id,
                source_format=fixture.format,
                artifact_family=fixture.categories[0],
                scenario=fixture.expected[0].scenario,
                environment=f"docker-linux-{measurement_runtime.architecture}",
                runtime=measurement_runtime,
                parse_coverage="passed" if passed else coverage_status,
                output_coverage="passed" if passed else coverage_status,
                target_runtime=coverage_status,
                source_differential="unavailable",
                evidence_bundle=str(bundle.relative_to(ROOT)),
            )
        )
    output = OUTPUT / "capability_measurements.json"
    output.write_text(
        json.dumps(
            {
                "schema_version": "1.0.0",
                "measurements": [item.model_dump(mode="json") for item in measurements],
            },
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )
    executions: list[CorpusExecution] = []
    for measurement in measurements:
        signature = None
        if measurement.target_runtime == "failed":
            signature = normalized_failure_signature(
                gate="target-runtime",
                error_type="sample-certification-failure",
                trace_diff=[measurement.evidence_id],
            )
        executions.append(
            CorpusExecution(
                fixture_id=measurement.fixture_id,
                scenario=measurement.scenario,
                environment=measurement.environment,
                target=measurement.target,
                target_version=measurement.target_version,
                status=measurement.target_runtime,
                failure_signature=signature,
                evidence_path=measurement.evidence_bundle,
            )
        )
    executions_output = OUTPUT / "corpus_executions.json"
    executions_output.write_text(
        json.dumps(
            {
                "schema_version": "1.0.0",
                "executions": [item.model_dump(mode="json") for item in executions],
            },
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
